"""
S-Curve Project Monitoring System — Streamlit Web App
Author: Faculty of Engineering, Chiang Mai University
Run:  streamlit run app.py
"""
import json, os, re, io, csv, sqlite3
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px

st.set_page_config(page_title="S-Curve Monitor", page_icon="📈",
                   layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
/* ── Base layout ── */
.stApp{background-color:#EFEFEF!important;color:#0A0A0A;font-family:'Inter',sans-serif}

/* ── Sidebar ── */
[data-testid="stSidebar"]{background-color:#FFFFFF!important;border-right:1px solid #E0E0E0}
[data-testid="stSidebar"] *{color:#0A0A0A!important}
/* Working project card — override sidebar wildcard */
[data-testid="stSidebar"] .proj-card *{color:inherit!important}
[data-testid="stSidebar"] .proj-label{color:#1AE06B!important}
[data-testid="stSidebar"] .proj-title{color:#FFFFFF!important}
[data-testid="stSidebar"] .proj-chip{color:#AAAAAA!important}
[data-testid="stSidebar"] .proj-chip-g{color:#1AE06B!important}

/* Nav radio — pill style */
[data-testid="stSidebar"] .stRadio [data-testid="stWidgetLabel"]{
  font-size:.7rem!important;font-weight:700!important;color:#9B9B9B!important;
  letter-spacing:1.5px;text-transform:uppercase;margin-bottom:6px}
[data-testid="stSidebar"] .stRadio [role="radiogroup"]{gap:1px!important;display:flex;flex-direction:column}
/* Each option row */
[data-testid="stSidebar"] .stRadio label{
  display:flex!important;align-items:center!important;width:100%!important;
  font-size:.88rem!important;font-weight:500!important;
  color:#6B6B6B!important;background:transparent!important;
  border-radius:10px!important;padding:9px 12px!important;
  cursor:pointer;transition:all .15s;margin:0!important}
[data-testid="stSidebar"] .stRadio label *{color:#6B6B6B!important}
/* Hide the actual radio circle */
[data-testid="stSidebar"] .stRadio input[type="radio"]{display:none!important}
[data-testid="stSidebar"] div[data-baseweb="radio"] > div:first-child{display:none!important}
/* Hover */
[data-testid="stSidebar"] .stRadio label:hover{
  background:#F0F0F0!important;color:#0A0A0A!important}
[data-testid="stSidebar"] .stRadio label:hover *{color:#0A0A0A!important}
/* Selected — :has() is supported Chrome 105+, Safari 15.4+, FF 121+ */
[data-testid="stSidebar"] .stRadio label:has(input:checked){
  background:#0A0A0A!important;color:#FFFFFF!important;font-weight:600!important;
  box-shadow:0 2px 8px rgba(0,0,0,.18)}
[data-testid="stSidebar"] .stRadio label:has(input:checked) *{color:#FFFFFF!important}

/* ── KPI cards ── */
.kpi-card{background:#FFFFFF;border-radius:16px;padding:20px 18px;text-align:left;
  margin:4px 0;box-shadow:0 2px 8px rgba(0,0,0,.06);transition:box-shadow .2s}
.kpi-card:hover{box-shadow:0 4px 16px rgba(0,0,0,.10)}
.kpi-num{font-size:2.4rem;font-weight:800;color:#0A0A0A;letter-spacing:-1px;line-height:1.1}
.kpi-num sup{font-size:1rem;font-weight:600;color:#6B6B6B}
.kpi-label{font-size:.72rem;color:#6B6B6B;letter-spacing:.5px;margin-top:6px;font-weight:500;text-transform:uppercase}
.kpi-green{color:#1AE06B!important}

/* ── Status badges ── */
.st-done{background:#E8FBF0;color:#0B9B41;padding:3px 12px;border-radius:20px;font-size:.8rem;font-weight:600}
.st-prog{background:#E8F8FF;color:#0070B8;padding:3px 12px;border-radius:20px;font-size:.8rem;font-weight:600}
.st-pend{background:#FFF8E6;color:#A06000;padding:3px 12px;border-radius:20px;font-size:.8rem;font-weight:600}
.st-delay{background:#FFF0F0;color:#CC2222;padding:3px 12px;border-radius:20px;font-size:.8rem;font-weight:600}
.st-none{background:#F2F2F2;color:#6B6B6B;padding:3px 12px;border-radius:20px;font-size:.8rem;font-weight:600}

/* ── Section headers ── */
.sec{font-size:.72rem;font-weight:700;color:#6B6B6B;letter-spacing:2px;text-transform:uppercase;
  border-bottom:1.5px solid #E0E0E0;padding-bottom:8px;margin:24px 0 14px 0}

/* ── Buttons ── */
.stButton>button{background:#0A0A0A!important;color:#FFFFFF!important;font-weight:600;
  border:none;border-radius:10px;padding:9px 22px;font-size:.88rem;transition:background .2s}
.stButton>button:hover{background:#1AE06B!important;color:#0A0A0A!important}
.stButton>button[kind="secondary"]{background:#FFFFFF!important;color:#0A0A0A!important;
  border:1.5px solid #E0E0E0!important}
.stButton>button[kind="secondary"]:hover{border-color:#1AE06B!important;color:#0B9B41!important}

/* ── Metrics ── */
div[data-testid="stMetricValue"]{color:#0A0A0A!important;font-size:1.7rem!important;font-weight:800!important}
div[data-testid="stMetricLabel"]{color:#6B6B6B!important;font-size:.75rem!important;text-transform:uppercase;letter-spacing:.5px}
div[data-testid="stMetricDelta"]{font-size:.8rem!important}

/* ── Widget labels (all variants) ── */
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span,
[data-testid="stWidgetLabel"] label,
label,
.stTextInput label, .stTextInput label p,
.stSelectbox label, .stSelectbox label p,
.stNumberInput label, .stNumberInput label p,
.stTextArea label, .stTextArea label p,
.stDateInput label, .stDateInput label p,
.stFileUploader label, .stFileUploader label p,
.stCheckbox label, .stCheckbox label p,
.stRadio label, .stRadio label p,
.stSlider label, .stSlider label p,
.stMultiSelect label, .stMultiSelect label p,
.stDateInput [data-testid="stWidgetLabel"] p,
div[class*="stForm"] label,
div[class*="stForm"] p{
  color:#0A0A0A!important;font-weight:600!important;font-size:.85rem!important}
/* form helper text / captions under inputs */
[data-testid="stWidgetLabel"]+div p,
[data-baseweb="form-control-label"]{color:#0A0A0A!important}

/* ── Expanders ── */
[data-testid="stExpander"]{background:#FFFFFF;border-radius:12px;border:1px solid #E0E0E0!important;margin-bottom:6px}
[data-testid="stExpander"] summary{font-weight:600!important;color:#0A0A0A!important}

/* ── Tabs ── */
[data-testid="stTabs"] [role="tablist"]{background:#EFEFEF;border-bottom:none!important;gap:4px;padding:4px}
[data-testid="stTabs"] [role="tab"]{
  font-weight:600;font-size:.85rem;color:#6B6B6B!important;
  background:transparent;border:none!important;border-radius:10px!important;
  padding:7px 18px!important;transition:all .18s}
[data-testid="stTabs"] [role="tab"] *,
[data-testid="stTabs"] [role="tab"] p,
[data-testid="stTabs"] [role="tab"] span{color:#6B6B6B!important}
[data-testid="stTabs"] [role="tab"]:hover{background:#E0E0E0!important;color:#0A0A0A!important}
[data-testid="stTabs"] [role="tab"]:hover *{color:#0A0A0A!important}
[data-testid="stTabs"] [role="tab"][aria-selected="true"]{
  background:#0A0A0A!important;color:#FFFFFF!important;
  border-radius:10px!important;border:none!important;
  box-shadow:0 2px 8px rgba(0,0,0,.18)}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] *,
[data-testid="stTabs"] [role="tab"][aria-selected="true"] p,
[data-testid="stTabs"] [role="tab"][aria-selected="true"] span{color:#FFFFFF!important}
[data-testid="stTabs"] [role="tab"][aria-selected="true"]:hover{background:#1AE06B!important;color:#0A0A0A!important}
[data-testid="stTabs"] [role="tab"][aria-selected="true"]:hover *{color:#0A0A0A!important}
[data-testid="stTabContent"]{padding-top:16px}

/* ── Data editor / tables ── */
.stDataFrame thead th{background:#F2F2F2!important;color:#0A0A0A!important;font-weight:700!important}
.stDataFrame{border-radius:12px;overflow:hidden}

/* ── Inputs ── */
input,textarea,select{color:#0A0A0A!important;background:#FFFFFF!important;border-radius:8px!important}
[data-testid="stSelectbox"] div,[data-testid="stMultiSelect"] div{color:#0A0A0A!important}

/* ── Text ── */
p,li,h1,h2,h3,h4,h5{color:#0A0A0A}
small,caption{color:#6B6B6B}
code{background:#F2F2F2;color:#0A0A0A;border-radius:4px;padding:1px 5px}

/* ── Info/success/warning boxes ── */
[data-testid="stAlert"]{border-radius:12px!important;border:none!important}
</style>""", unsafe_allow_html=True)

# ── Project Database (SQLite) ─────────────────────────────────────────────────
DB_FILE = "projects.db"

def db_connect():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def db_init():
    """Create tables and migrate from manifest JSON if needed."""
    conn = db_connect()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            key         TEXT    UNIQUE,
            project_name     TEXT,
            project_name_th  TEXT,
            doc         TEXT,
            phase       TEXT,
            n_activities     INTEGER DEFAULT 0,
            total_budget     REAL    DEFAULT 0,
            data_json   TEXT,
            created_at  TEXT,
            updated_at  TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS save_log (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            db_key       TEXT,
            project_name TEXT,
            action       TEXT,
            detail       TEXT,
            timestamp    TEXT
        )
    """)
    conn.commit()
    # Migrate from project_manifest.json + JSON files if DB is empty
    if conn.execute("SELECT COUNT(*) FROM projects").fetchone()[0] == 0:
        mf = "projects/project_manifest.json"
        if os.path.exists(mf):
            with open(mf, encoding="utf-8") as f:
                manifest = json.load(f)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            for e in manifest:
                fpath = os.path.join("projects", e["filename"])
                d_json = ""
                if os.path.exists(fpath):
                    with open(fpath, encoding="utf-8") as pf:
                        d_json = pf.read()
                try:
                    conn.execute("""
                        INSERT OR IGNORE INTO projects
                        (key,project_name,project_name_th,doc,phase,n_activities,total_budget,data_json,created_at,updated_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, (e["key"], e["project_name"], e.get("project_name_th",""),
                          e["doc"], e.get("phase",""), e["n_activities"],
                          e.get("total_budget",0), d_json, now, now))
                except Exception:
                    pass
            conn.commit()
    conn.close()

def db_list(sort="updated_at", search=""):
    """Return all projects sorted by updated_at DESC."""
    conn = db_connect()
    q = "SELECT id,key,project_name,project_name_th,doc,phase,n_activities,total_budget,created_at,updated_at FROM projects"
    params = []
    if search:
        q += " WHERE project_name LIKE ? OR doc LIKE ? OR phase LIKE ?"
        params = [f"%{search}%"]*3
    q += f" ORDER BY {sort} DESC"
    rows = [dict(r) for r in conn.execute(q, params).fetchall()]
    conn.close()
    return rows

def db_get(key):
    conn = db_connect()
    r = conn.execute("SELECT data_json FROM projects WHERE key=?", (key,)).fetchone()
    conn.close()
    return json.loads(r["data_json"]) if r and r["data_json"] else None

def db_save(key, project_name, project_name_th, doc, phase, proj_data):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    n = len(proj_data.get("activities", []))
    b = proj_data.get("total_budget", 0)
    j = json.dumps(proj_data, ensure_ascii=False)
    conn = db_connect()
    conn.execute("""
        INSERT INTO projects (key,project_name,project_name_th,doc,phase,n_activities,total_budget,data_json,created_at,updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(key) DO UPDATE SET
            project_name=excluded.project_name,
            project_name_th=excluded.project_name_th,
            doc=excluded.doc, phase=excluded.phase,
            n_activities=excluded.n_activities,
            total_budget=excluded.total_budget,
            data_json=excluded.data_json,
            updated_at=excluded.updated_at
    """, (key, project_name, project_name_th, doc, phase, n, b, j, now, now))
    conn.commit()
    conn.close()

def db_touch(key, project_name=""):
    """Update updated_at timestamp (on Load)."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = db_connect()
    if not project_name:
        r = conn.execute("SELECT project_name FROM projects WHERE key=?", (key,)).fetchone()
        project_name = r["project_name"] if r else ""
    conn.execute("UPDATE projects SET updated_at=? WHERE key=?", (now, key))
    conn.commit(); conn.close()
    db_log(key, project_name, "▶ Loaded", "Project loaded into active workspace")

def db_delete(key):
    conn = db_connect()
    conn.execute("DELETE FROM projects WHERE key=?", (key,))
    conn.commit(); conn.close()

def db_log(db_key, project_name, action, detail=""):
    """Write one entry to the activity log."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = db_connect()
    conn.execute(
        "INSERT INTO save_log (db_key,project_name,action,detail,timestamp) VALUES (?,?,?,?,?)",
        (db_key or "", project_name or "", action, detail, now))
    conn.commit(); conn.close()

def db_get_log(limit=50, db_key=None):
    """Return recent log entries, optionally filtered to one project."""
    conn = db_connect()
    if db_key:
        rows = conn.execute(
            "SELECT * FROM save_log WHERE db_key=? ORDER BY timestamp DESC LIMIT ?",
            (db_key, limit)).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM save_log ORDER BY timestamp DESC LIMIT ?",
            (limit,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

db_init()   # run once on startup

# ── Data ─────────────────────────────────────────────────────────────────────
DATA_FILE = "project_data.json"

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, encoding="utf-8") as f:
            d = json.load(f)
        # Ensure _db_key is always present after loading from disk
        if not d.get("_db_key"):
            raw = (d.get("project_name","project") + "_" + d.get("start_date","")[:7])
            d["_db_key"]   = re.sub(r'[^A-Za-z0-9_]', '_', raw)[:40]
            d["_db_doc"]   = d.get("_db_doc", "CUSTOM")
            d["_db_phase"] = d.get("_db_phase", "Custom")
            # Write key back so it persists
            with open(DATA_FILE,"w",encoding="utf-8") as f:
                json.dump(d,f,ensure_ascii=False,indent=2)
        return d
    return {"project_name":"Demo Project","project_name_th":"โครงการตัวอย่าง",
            "contract_no":"SF-NRCT-FY2569","project_owner":"NRCT",
            "contractor":"Faculty of Engineering, CMU","start_date":"2025-10-01",
            "end_date":"2026-09-30","total_budget":8166000,"n_months":12,"activities":[]}

def save_data(d):
    # ── Auto-stamp _db_key if missing so EVERY save syncs to SQLite ──────────
    if not d.get("_db_key"):
        raw = (d.get("project_name","project") + "_" + d.get("start_date","")[:7])
        d["_db_key"]   = re.sub(r'[^A-Za-z0-9_]', '_', raw)[:40]
        d["_db_doc"]   = d.get("_db_doc", "CUSTOM")
        d["_db_phase"] = d.get("_db_phase", "Custom")

    with open(DATA_FILE,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False,indent=2)
    st.cache_data.clear()

    # Always sync to SQLite
    _key = d["_db_key"]
    n = len(d.get("activities",[]))
    db_save(_key, d.get("project_name",""), d.get("project_name_th",""),
            d.get("_db_doc","CUSTOM"), d.get("_db_phase","Custom"), d)
    db_log(_key, d.get("project_name",""), "💾 Saved",
           f"{n} activities | budget ฿{d.get('total_budget',0):,.0f}")

@st.cache_data
def get_labels(start_iso, n):
    base=datetime.strptime(start_iso,"%Y-%m-%d")
    return [(base+relativedelta(months=m)).strftime("%b %Y") for m in range(n)]

def cur_month(start_iso,n):
    base=datetime.strptime(start_iso,"%Y-%m-%d"); t=datetime.today()
    return max(1,min((t.year-base.year)*12+(t.month-base.month)+1,n))

def compute(data):
    n=data["n_months"]; pm=[0.0]*n; am=[0.0]*n
    for a in data["activities"]:
        s,e=int(a["start_month"]),int(a["end_month"]); dur=max(1,e-s+1); mp=a["weight"]/dur
        for m in range(s,min(e+1,n+1)):
            pm[m-1]+=mp
            v=a.get("actuals",{}).get(str(m))
            if v is not None: am[m-1]+=float(v)*a["weight"]/100/dur
    cp=ca=0.0; cump=[]; cuma=[]
    for i in range(n):
        cp+=pm[i]; cump.append(round(cp,2))
        ca+=am[i]; cuma.append(round(ca,2) if (am[i]>0 or ca>0) else None)
    return cump,cuma

def compute_cost(data):
    """Cumulative planned-cost and actual-cost arrays by month (฿)."""
    acts=data["activities"]; N=data["n_months"]
    budget=data.get("total_budget",0)
    mp=[0.0]*N; ma=[None]*N
    for a in acts:
        s,e=int(a["start_month"]),int(a["end_month"]); dur=max(1,e-s+1)
        cost=a.get("planned_cost",0) or round(a.get("weight",0)/100*budget,2)
        per=cost/dur
        for m in range(s,min(e+1,N+1)):
            mp[m-1]+=per
        for m_str,v in a.get("actual_costs",{}).items():
            m=int(m_str)
            if 1<=m<=N:
                ma[m-1]=(ma[m-1] or 0)+float(v)
    cum_p=cum_a=0.0; cp=[]; ca=[]
    for i in range(N):
        cum_p+=mp[i]; cp.append(round(cum_p,2))
        if ma[i] is not None:
            cum_a+=ma[i]; ca.append(round(cum_a,2))
        else:
            ca.append(round(cum_a,2) if cum_a>0 else None)
    return cp,ca,mp,ma

def build_excel(data, cump, cuma, labels):
    """Build a formatted .xlsx workbook and return bytes."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    import io as _io

    wb = openpyxl.Workbook()
    acts = data.get("activities", [])
    N    = data.get("n_months", 12)

    # ── Colour palette ────────────────────────────────────────────────────────
    HDR_BG  = PatternFill("solid", fgColor="2D3747")
    HDR_FG  = Font(color="E8ECF0", bold=True, size=10)
    LOCK_BG = PatternFill("solid", fgColor="3D4557")
    EDIT_BG = PatternFill("solid", fgColor="1A2030")
    CUR_BG  = PatternFill("solid", fgColor="5A3A00")   # orange tint = current month
    DONE_BG = PatternFill("solid", fgColor="2A5E3A")
    PROG_BG = PatternFill("solid", fgColor="2A4E2A")
    DEL_BG  = PatternFill("solid", fgColor="5A1A1A")
    THIN    = Border(
        left=Side(style="thin", color="555B6E"),
        right=Side(style="thin", color="555B6E"),
        top=Side(style="thin", color="555B6E"),
        bottom=Side(style="thin", color="555B6E"))
    WH  = Font(color="E8ECF0", size=9)
    GRY = Font(color="9BA3AF", size=9)

    def _ap(ws, r, c, v, fill=None, font=None, align=None, num_fmt=None):
        cell = ws.cell(row=r, column=c, value=v)
        if fill:    cell.fill   = fill
        if font:    cell.font   = font
        if align:   cell.alignment = align
        if num_fmt: cell.number_format = num_fmt
        cell.border = THIN
        return cell

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # ── Sheet 1 : Progress (editable) ────────────────────────────────────────
    ws = wb.active
    ws.title = "Progress"
    ws.freeze_panes = "F2"          # freeze cols A-E, row 1
    ws.sheet_view.showGridLines = False

    # Header row
    fixed_hdrs = ["No", "Activity", "Wt%", "Cost (฿)", "Status"]
    month_hdrs = [labels[m] if m < len(labels) else f"M{m+1}" for m in range(N)]
    all_hdrs   = fixed_hdrs + month_hdrs
    cm_idx     = cur_month(data["start_date"], N)   # 1-based

    for ci, h in enumerate(all_hdrs, 1):
        is_cm = (ci == 5 + cm_idx)
        bg = CUR_BG if is_cm else HDR_BG
        _ap(ws, 1, ci, h, fill=bg, font=HDR_FG, align=center)

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 7
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    for ci in range(6, 6+N):
        ws.column_dimensions[get_column_letter(ci)].width = 9

    # Status dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Pending,Completed,Delayed,Cancelled"',
        showDropDown=False)
    ws.add_data_validation(dv)

    sopts = ["Not Started","In Progress","Pending","Completed","Delayed","Cancelled"]

    for ri, a in enumerate(acts, 2):
        sk = next((k for k in sopts if k in a.get("status","")), sopts[0])
        # Status row colour
        if   "Completed"   in sk: row_bg = DONE_BG
        elif "In Progress" in sk: row_bg = PROG_BG
        elif "Delayed"     in sk: row_bg = DEL_BG
        else:                     row_bg = EDIT_BG

        _ap(ws, ri, 1, a.get("no",""),  fill=LOCK_BG, font=GRY,  align=center)
        _ap(ws, ri, 2, a.get("name",""),fill=LOCK_BG, font=WH,   align=left)
        _ap(ws, ri, 3, a.get("weight",0),fill=LOCK_BG,font=GRY,  align=center, num_fmt="0.0")
        _ap(ws, ri, 4, a.get("planned_cost",0), fill=LOCK_BG, font=GRY, align=center, num_fmt='#,##0')
        _ap(ws, ri, 5, sk,               fill=row_bg, font=WH,   align=center)
        dv.add(ws.cell(row=ri, column=5))

        for mi in range(1, N+1):
            ci = 5 + mi
            in_range = a["start_month"] <= mi <= a["end_month"]
            val = float(a.get("actuals",{}).get(str(mi), 0)) if in_range else None
            is_cm_col = (mi == cm_idx)
            bg = CUR_BG if is_cm_col else (EDIT_BG if in_range else LOCK_BG)
            fn = WH if in_range else GRY
            cell = _ap(ws, ri, ci, val, fill=bg, font=fn, align=center, num_fmt="0")
            if not in_range:
                cell.protection = openpyxl.styles.Protection(locked=True)

        ws.row_dimensions[ri].height = 18

    ws.row_dimensions[1].height = 22

    # ── Sheet 2 : S-Curve Data (reference) ──────────────────────────────────
    ws2 = wb.create_sheet("S-Curve Data")
    ws2.freeze_panes = "B2"
    ws2.sheet_view.showGridLines = False
    for ci, h in enumerate(["Month","Plan (%)","Actual (%)","SV","SPI"], 1):
        _ap(ws2, 1, ci, h, fill=HDR_BG, font=HDR_FG, align=center)
    ws2.column_dimensions["A"].width = 14
    for col in ["B","C","D","E"]: ws2.column_dimensions[col].width = 12
    for i, lbl in enumerate(labels):
        pv = cump[i]
        ev = cuma[i]
        sv_ = round(ev-pv,2) if ev is not None else None
        spi_= round(ev/pv,2) if (ev is not None and pv>0) else None
        row = [lbl, pv, ev, sv_, spi_]
        for ci, v in enumerate(row, 1):
            _ap(ws2, i+2, ci, v, fill=EDIT_BG, font=WH, align=center, num_fmt="0.00")

    # ── Sheet 3 : Project Info ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Project Info")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 50
    fields = [
        ("Project Name (EN)", data.get("project_name","")),
        ("Project Name (TH)", data.get("project_name_th","")),
        ("Contract No.",      data.get("contract_no","")),
        ("Project Owner",     data.get("project_owner","")),
        ("Contractor",        data.get("contractor","")),
        ("Start Date",        data.get("start_date","")),
        ("End Date",          data.get("end_date","")),
        ("Months",            data.get("n_months","")),
        ("Total Budget (THB)",data.get("total_budget",0)),
        ("No. Activities",    len(acts)),
    ]
    for ri, (k, v) in enumerate(fields, 1):
        _ap(ws3, ri, 1, k, fill=HDR_BG, font=HDR_FG, align=left)
        _ap(ws3, ri, 2, v, fill=EDIT_BG, font=WH,    align=left)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

SCLS={"Completed":"st-done","In Progress":"st-prog","Pending":"st-pend",
      "Delayed":"st-delay","Not Started":"st-none"}
SMAP={"Completed":"✅ Completed","In Progress":"🚧 In Progress","Pending":"⏳ Pending",
      "Delayed":"💤 Delayed","Not Started":"❌ Not Started","Cancelled":"🚫 Cancelled"}
SCOL={"Completed":"#1AE06B","In Progress":"#0A0A0A","Pending":"#A06000",
      "Delayed":"#CC2222","Not Started":"#C0C0C0"}

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    # ── App title ──────────────────────────────────────────────────────────────
    st.markdown(
        "<div style='font-size:1.1rem;font-weight:800;color:#0A0A0A;letter-spacing:-.3px;margin-bottom:2px'>"
        "📈 S-Curve Monitor</div>"
        "<div style='font-size:.72rem;color:#6B6B6B;margin-bottom:14px'>Project Performance Dashboard</div>",
        unsafe_allow_html=True)

    # ── Active project card (top of nav) ───────────────────────────────────────
    _data_sb = load_data()
    _aname   = _data_sb.get("project_name", "No project loaded")
    _cm_sb   = cur_month(_data_sb["start_date"], _data_sb["n_months"])
    _phase   = _data_sb.get("_db_phase", "")
    _n_acts  = len(_data_sb.get("activities", []))
    st.markdown(
        f"<div class='proj-card' style='background:#0A0A0A;border-radius:14px;padding:12px 14px;margin-bottom:16px'>"
        f"<div class='proj-label' style='font-size:.62rem;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:6px'>▶ Working Project</div>"
        f"<div class='proj-title' style='font-size:.88rem;font-weight:700;line-height:1.3;margin-bottom:8px'>{_aname[:42]}</div>"
        f"<div style='display:flex;gap:6px;flex-wrap:wrap'>"
        f"<span class='proj-chip' style='background:#1A1A1A;font-size:.68rem;padding:2px 8px;border-radius:6px'>📁 {_phase or 'General'}</span>"
        f"<span class='proj-chip' style='background:#1A1A1A;font-size:.68rem;padding:2px 8px;border-radius:6px'>📋 {_n_acts} tasks</span>"
        f"<span class='proj-chip-g' style='background:#1AE06B22;font-size:.68rem;padding:2px 8px;border-radius:6px;font-weight:700'>M{_cm_sb} now</span>"
        f"</div></div>",
        unsafe_allow_html=True)

    # ── Navigation ─────────────────────────────────────────────────────────────
    st.markdown("<div style='font-size:.68rem;font-weight:700;color:#6B6B6B;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:6px'>NAVIGATE</div>", unsafe_allow_html=True)
    page=st.radio("Navigate",[
        "🏠 Dashboard",
        "📚 WBS Library",
        "⚙️ Project Setup",
        "📝 Update Progress",
        "📈 S-Curve",
        "💰 Cost Tracking",
        "📅 Gantt",
        "📊 EVM Indicators",
    ], label_visibility="collapsed")

    # ── Footer info ────────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown(
        f"<div style='font-size:.75rem;color:#6B6B6B'>"
        f"📅 <b style='color:#0A0A0A'>M{_cm_sb}</b> &nbsp;·&nbsp; "
        f"🗓️ <b style='color:#0A0A0A'>{date.today().strftime('%d %b %Y')}</b></div>",
        unsafe_allow_html=True)
    st.caption("Ping River Basin • Chiang Mai University © 2026")

data=load_data()
acts=data["activities"]; N=data["n_months"]; SI=data["start_date"]
labels=get_labels(SI,N)
cump,cuma=compute(data)
cum_pc,cum_ac,monthly_pc,monthly_ac=compute_cost(data)
cm=cur_month(SI,N)

# ── DASHBOARD ─────────────────────────────────────────────────────────────────
if page=="🏠 Dashboard":
    st.markdown(f"# 🏠 Dashboard")
    st.markdown(f"**{data['project_name']}**  \n<small>{data.get('project_name_th','')}</small>",unsafe_allow_html=True)
    st.markdown("---")

    total=len(acts)
    done=sum(1 for a in acts if "Completed" in a.get("status",""))
    prog=sum(1 for a in acts if "In Progress" in a.get("status",""))
    delay=sum(1 for a in acts if "Delayed" in a.get("status",""))
    pv=cump[cm-1] if cm<=N else cump[-1]
    ev=next((cuma[i] for i in range(cm-1,-1,-1) if cuma[i] is not None),0.0) or 0.0
    spi=round(ev/pv,2) if pv>0 else None; sv=round(ev-pv,2) if pv>0 else None

    c1,c2,c3,c4,c5,c6=st.columns(6)
    cards=[
        (c1,str(total),"Total Tasks","●","#0A0A0A"),
        (c2,str(done),"Completed","●","#1AE06B"),
        (c3,str(prog),"In Progress","●","#0070B8"),
        (c4,str(delay),"Delayed","●","#CC2222"),
        (c5,f"{spi:.2f}" if spi else "—","SPI Index","●",
         "#1AE06B" if spi and spi>=1 else "#A06000" if spi and spi>=0.8 else "#CC2222"),
        (c6,f"{sv:+.1f}%" if sv is not None else "—","Schedule Var.","●",
         "#1AE06B" if sv and sv>=0 else "#CC2222"),
    ]
    for col,val,lbl,dot,ac in cards:
        col.markdown(
            f'<div class="kpi-card">'
            f'<div class="kpi-num">{val}<sup style="font-size:.9rem;color:{ac}"> {dot}</sup></div>'
            f'<div class="kpi-label">{lbl}</div>'
            f'</div>', unsafe_allow_html=True)

    # ── Cost KPI row ──────────────────────────────────────────────────────────
    budget_d = data.get("total_budget", 0)
    total_wbs_d = sum(a.get("planned_cost", 0) or round(a.get("weight", 0) / 100 * budget_d, 2) for a in acts)
    total_actual_cost_d = next((v for v in reversed(cum_ac) if v is not None), 0) or 0
    remaining_d = budget_d - total_actual_cost_d
    pv_cost_d = cum_pc[cm - 1] if cm <= N else cum_pc[-1]
    cost_var_d = total_actual_cost_d - pv_cost_d

    st.markdown("")
    cc1, cc2, cc3, cc4, cc5 = st.columns(5)
    cost_cards = [
        (cc1, f"฿{budget_d:,.0f}", "Contract Budget", "#0A0A0A"),
        (cc2, f"฿{total_wbs_d:,.0f}", "WBS Planned Cost",
         "#1AE06B" if total_wbs_d <= budget_d else "#CC2222"),
        (cc3, f"฿{total_actual_cost_d:,.0f}", f"Actual Cost (M{cm})",
         "#1AE06B" if total_actual_cost_d <= pv_cost_d else "#CC2222"),
        (cc4, f"฿{remaining_d:,.0f}", "Budget Remaining",
         "#1AE06B" if remaining_d >= 0 else "#CC2222"),
        (cc5, f"฿{cost_var_d:+,.0f}", "Cost Variance",
         "#1AE06B" if cost_var_d <= 0 else "#CC2222"),
    ]
    for col, val, lbl, ac in cost_cards:
        col.markdown(
            f'<div class="kpi-card">'
            f'<div class="kpi-num" style="color:{ac};font-size:1.6rem">{val}</div>'
            f'<div class="kpi-label">{lbl}</div>'
            f'</div>', unsafe_allow_html=True)

    st.markdown("")
    cl,cr=st.columns([2,1])
    with cl:
        st.markdown('<div class="sec">S - C U R V E    O V E R V I E W</div>',unsafe_allow_html=True)
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=labels,y=cump,name="Plan",fill="tozeroy",
            fillcolor="rgba(26,224,107,.12)",line=dict(color="#0A0A0A",width=2.5),
            mode="lines+markers",marker=dict(size=7),
            hovertemplate="<b>%{x}</b><br>Plan: %{y:.1f}%<extra></extra>"))
        ax=[labels[i] for i,v in enumerate(cuma) if v is not None]
        ay=[v for v in cuma if v is not None]
        if ay:
            fig.add_trace(go.Scatter(x=ax,y=ay,name="Actual",fill="tozeroy",
                fillcolor="rgba(26,224,107,.18)",line=dict(color="#1AE06B",width=2.5),
                mode="lines+markers",marker=dict(size=8,symbol="diamond"),
                hovertemplate="<b>%{x}</b><br>Actual: %{y:.1f}%<extra></extra>"))
        if cm<=N:
            fig.add_vline(x=labels[cm-1],line_dash="dash",line_color="#1AE06B",line_width=2)
            fig.add_annotation(x=labels[cm-1],y=105,text="TODAY",showarrow=False,
                font=dict(color="#0A0A0A",size=11,family="sans-serif"),xanchor="left")
        fig.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
            font_color="#0A0A0A",height=280,margin=dict(l=10,r=10,t=10,b=10),
            xaxis=dict(gridcolor="#F0F0F0",showline=False),yaxis=dict(gridcolor="#F0F0F0",range=[0,105]),
            legend=dict(bgcolor="#FFFFFF",font=dict(color="#0A0A0A")),hovermode="x unified")
        st.plotly_chart(fig,use_container_width=True)

    with cr:
        st.markdown('<div class="sec">A C T I V I T I E S</div>',unsafe_allow_html=True)
        for a in acts:
            st_raw = a.get("status","Not Started")
            sk = next((k for k in SCLS if k in st_raw),"Not Started")
            is_active = a["start_month"] <= cm <= a["end_month"]
            # Highlight border for currently active activities
            border_style = "border-left:3px solid #1AE06B;padding-left:6px;" if is_active else "border-left:3px solid #E8E8E8;padding-left:6px;"
            active_dot = '<span style="color:#1AE06B;font-size:.7rem;margin-right:3px">●</span>' if is_active else ''
            st.markdown(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:5px 0;border-bottom:1px solid #F0F0F0;{border_style}">'
                f'<span style="color:#0A0A0A;font-size:.82rem">{active_dot}{a["no"]}. {a["name"][:32]}…</span>'
                f'<span class="{SCLS[sk]}">{sk}</span></div>',
                unsafe_allow_html=True)

    # ── ACTIVE THIS MONTH spotlight ───────────────────────────────────────────
    active_acts = [a for a in acts if a["start_month"] <= cm <= a["end_month"]]
    if active_acts:
        st.markdown(f'<div class="sec">⏰ A C T I V E &nbsp; T H I S &nbsp; M O N T H &nbsp;— M{cm} &nbsp;({labels[cm-1] if cm<=N else ""})</div>',
                    unsafe_allow_html=True)
        cols_h = st.columns(min(len(active_acts), 3))
        for idx, a in enumerate(active_acts):
            st_raw = a.get("status","Not Started")
            sk = next((k for k in SCLS if k in st_raw),"Not Started")
            actual_cm = float(a.get("actuals",{}).get(str(cm), 0))
            # Health color
            if "Completed" in st_raw:    ac_col, hlbl = "#1AE06B", "✅ Done"
            elif "Delayed" in st_raw:    ac_col, hlbl = "#CC2222", "⚠️ Delayed"
            elif actual_cm > 0:          ac_col, hlbl = "#0A0A0A", "🚧 Active"
            else:                        ac_col, hlbl = "#A06000", "⏳ Needs Input"
            col = cols_h[idx % 3]
            col.markdown(f"""
<div style="background:#FFFFFF;border:none;border-radius:16px;padding:14px 16px;margin:4px 0;box-shadow:0 2px 8px rgba(0,0,0,.07)">
  <div style="display:flex;justify-content:space-between;align-items:center">
    <span style="color:#6B6B6B;font-size:.68rem;font-weight:700;letter-spacing:1.5px;text-transform:uppercase">Activity {a['no']} · Wt {a['weight']}%</span>
    <span style="background:{ac_col};color:#fff;font-size:.68rem;font-weight:700;padding:2px 9px;border-radius:20px">{hlbl}</span>
  </div>
  <div style="color:#0A0A0A;font-size:.9rem;font-weight:700;margin:8px 0 4px">{a['name'][:50]}</div>
  <div style="color:#6B6B6B;font-size:.75rem;margin-bottom:8px">M{a['start_month']} → M{a['end_month']}</div>
  <div style="background:#EFEFEF;border-radius:6px;height:5px">
    <div style="background:{ac_col};height:5px;border-radius:6px;width:{min(actual_cm,100):.0f}%"></div>
  </div>
  <div style="color:#6B6B6B;font-size:.7rem;margin-top:4px">M{cm} actual: {actual_cm:.0f}%</div>
</div>""", unsafe_allow_html=True)

    st.markdown("---")
    p1,p2,p3,p4,p5=st.columns(5)
    p1.metric("Contract",data.get("contract_no","—"))
    p2.metric("Start",data["start_date"]); p3.metric("End",data["end_date"])
    p4.metric("Budget","฿{:,.0f}".format(data["total_budget"]))
    total_planned_cost = sum(a.get("planned_cost",0) for a in acts)
    p5.metric("WBS Cost","฿{:,.0f}".format(total_planned_cost),
              delta=f"฿{total_planned_cost - data['total_budget']:+,.0f}" if total_planned_cost else None,
              delta_color="inverse")

# ── S-CURVE ───────────────────────────────────────────────────────────────────
elif page=="📈 S-Curve":
    st.markdown("# 📈 S-Curve")
    st.markdown("---")
    sc_tab1, sc_tab2, sc_tab3 = st.tabs(["📈 Progress S-Curve", "💰 Cost S-Curve", "🌐 Integrated Plan"])

    # ── Tab 1: Progress ────────────────────────────────────────────────────────
    with sc_tab1:
        fig=go.Figure()
        fig.add_trace(go.Scatter(x=labels,y=cump,name="Planned (%)",fill="tozeroy",
            fillcolor="rgba(26,224,107,.12)",line=dict(color="#0A0A0A",width=2.5),
            mode="lines+markers",marker=dict(size=8),
            hovertemplate="<b>%{x}</b><br>Plan: %{y:.1f}%<extra></extra>"))
        ax=[labels[i] for i,v in enumerate(cuma) if v is not None]
        ay=[v for v in cuma if v is not None]
        if ay:
            fig.add_trace(go.Scatter(x=ax,y=ay,name="Actual (%)",fill="tozeroy",
                fillcolor="rgba(26,224,107,.18)",line=dict(color="#1AE06B",width=2.5),
                mode="lines+markers",marker=dict(size=9,symbol="diamond"),
                hovertemplate="<b>%{x}</b><br>Actual: %{y:.1f}%<extra></extra>"))
        if cm<=N:
            fig.add_vline(x=labels[cm-1],line_dash="dash",line_color="#1AE06B",line_width=2)
            fig.add_annotation(x=labels[cm-1],y=105,text=f"M{cm} TODAY",showarrow=False,
                font=dict(color="#0A0A0A",size=11),xanchor="left")
        fig.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A",size=13),height=420,margin=dict(l=20,r=20,t=30,b=20),
            xaxis=dict(title="Month",gridcolor="#F0F0F0",tickangle=-30,showline=False),
            yaxis=dict(title="Cumulative Progress (%)",gridcolor="#F0F0F0",range=[0,105],ticksuffix="%"),
            legend=dict(bgcolor="#FFFFFF",bordercolor="#E0E0E0",borderwidth=1),hovermode="x unified")
        st.plotly_chart(fig,use_container_width=True)

        rows=[]
        for i,lbl in enumerate(labels):
            pv=cump[i]; ev=cuma[i]
            sv_=round(ev-pv,2) if ev is not None else None
            spi_=round(ev/pv,2) if (ev is not None and pv>0) else None
            rows.append({"Month":lbl,"Plan (%)":f"{pv:.1f}",
                "Actual (%)":f"{ev:.1f}" if ev is not None else "—",
                "SV":f"{sv_:+.1f}" if sv_ is not None else "—",
                "SPI":f"{spi_:.2f}" if spi_ else "—",
                "Status":("✅ On Track" if spi_ and spi_>=1.05 else "⚠️ Warning" if spi_ and spi_>=0.95
                          else "💤 Delayed" if spi_ and spi_>=0.8 else "🔴 Critical" if spi_ else "—")})
        st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

    # ── Tab 2: Cost S-Curve ────────────────────────────────────────────────────
    with sc_tab2:
        budget = data.get("total_budget",0)
        total_wbs = sum(a.get("planned_cost",0) or round(a.get("weight",0)/100*budget,2) for a in acts)
        burn_down = [budget - v for v in cum_pc]

        # KPI row
        total_actual_cost = next((v for v in reversed(cum_ac) if v is not None), 0) or 0
        remaining = budget - total_actual_cost
        cost_var  = total_actual_cost - (cum_pc[cm-1] if cm<=N else cum_pc[-1])

        ck1,ck2,ck3,ck4 = st.columns(4)
        def _ck(col,val,lbl,ac="#0A0A0A"):
            col.markdown(f'<div class="kpi-card"><div class="kpi-num" style="color:{ac}">{val}</div>'
                         f'<div class="kpi-label">{lbl}</div></div>',unsafe_allow_html=True)
        _ck(ck1,f"฿{budget:,.0f}","Contract Budget")
        _ck(ck2,f"฿{total_wbs:,.0f}","WBS Planned Cost",
            "#1AE06B" if total_wbs<=budget else "#CC2222")
        _ck(ck3,f"฿{total_actual_cost:,.0f}",f"Actual Cost (M{cm})",
            "#1AE06B" if total_actual_cost<=(cum_pc[cm-1] if cm<=N else 0) else "#CC2222")
        _ck(ck4,f"฿{remaining:,.0f}","Budget Remaining",
            "#1AE06B" if remaining>=0 else "#CC2222")

        st.markdown("")

        # Cost S-Curve chart
        fig2=go.Figure()
        fig2.add_trace(go.Scatter(x=labels,y=cum_pc,name="Planned Cost (฿)",
            fill="tozeroy",fillcolor="rgba(10,10,10,.07)",
            line=dict(color="#0A0A0A",width=2.5),mode="lines+markers",marker=dict(size=7),
            hovertemplate="<b>%{x}</b><br>Plan: ฿%{y:,.0f}<extra></extra>"))
        ax2=[labels[i] for i,v in enumerate(cum_ac) if v is not None]
        ay2=[v for v in cum_ac if v is not None]
        if ay2:
            fig2.add_trace(go.Scatter(x=ax2,y=ay2,name="Actual Cost (฿)",
                fill="tozeroy",fillcolor="rgba(26,224,107,.15)",
                line=dict(color="#1AE06B",width=2.5),mode="lines+markers",
                marker=dict(size=9,symbol="diamond"),
                hovertemplate="<b>%{x}</b><br>Actual: ฿%{y:,.0f}<extra></extra>"))
        fig2.add_trace(go.Scatter(x=labels,y=burn_down,name="Budget Remaining (฿)",
            line=dict(color="#CC2222",width=1.5,dash="dot"),
            mode="lines",yaxis="y2",
            hovertemplate="<b>%{x}</b><br>Remaining: ฿%{y:,.0f}<extra></extra>"))
        if cm<=N:
            fig2.add_vline(x=labels[cm-1],line_dash="dash",line_color="#1AE06B",line_width=2)
            fig2.add_annotation(x=labels[cm-1],y=budget,text="TODAY",showarrow=False,
                font=dict(color="#0A0A0A",size=11),xanchor="left")
        fig2.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A",size=13),height=400,margin=dict(l=20,r=60,t=30,b=20),
            xaxis=dict(title="Month",gridcolor="#F0F0F0",tickangle=-30,showline=False),
            yaxis=dict(title="Cumulative Cost (฿)",gridcolor="#F0F0F0",tickprefix="฿",tickformat=",.0f"),
            yaxis2=dict(title="Remaining (฿)",overlaying="y",side="right",
                        tickprefix="฿",tickformat=",.0f",showgrid=False),
            legend=dict(bgcolor="#FFFFFF",bordercolor="#E0E0E0",borderwidth=1),
            hovermode="x unified")
        st.plotly_chart(fig2,use_container_width=True)

        # Monthly bar
        st.markdown('<div class="sec">MONTHLY PLANNED SPEND</div>',unsafe_allow_html=True)
        act_colors=["#1AE06B" if (monthly_ac[i] or 0)<=monthly_pc[i] else "#CC2222" for i in range(N)]
        fig3=go.Figure()
        fig3.add_trace(go.Bar(x=labels,y=monthly_pc,name="Planned",
            marker_color="#0A0A0A",opacity=0.85,
            hovertemplate="<b>%{x}</b><br>Plan: ฿%{y:,.0f}<extra></extra>"))
        ma_vals=[monthly_ac[i] or 0 for i in range(N)]
        if any(v>0 for v in ma_vals):
            fig3.add_trace(go.Bar(x=labels,y=ma_vals,name="Actual",
                marker_color="#1AE06B",opacity=0.9,
                hovertemplate="<b>%{x}</b><br>Actual: ฿%{y:,.0f}<extra></extra>"))
        fig3.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A",size=12),height=300,barmode="group",
            margin=dict(l=20,r=20,t=20,b=20),
            xaxis=dict(gridcolor="#F0F0F0",showline=False,tickangle=-30),
            yaxis=dict(gridcolor="#F0F0F0",tickprefix="฿",tickformat=",.0f"),
            legend=dict(bgcolor="#FFFFFF"),hovermode="x unified")
        st.plotly_chart(fig3,use_container_width=True)

        # Cost detail table
        cost_rows=[]
        for a in acts:
            pc=a.get("planned_cost",0) or round(a.get("weight",0)/100*budget,2)
            ac_tot=sum(float(v) for v in a.get("actual_costs",{}).values())
            share=round(pc/total_wbs*100,1) if total_wbs>0 else 0
            cost_rows.append({"No":a["no"],"Activity":a["name"][:45],
                "Wt%":a["weight"],"Budget(฿)":f"฿{pc:,.0f}",
                "Actual(฿)":f"฿{ac_tot:,.0f}","Remaining(฿)":f"฿{pc-ac_tot:,.0f}",
                "Share%":f"{share}%","Status":a.get("status","Not Started")})
        st.dataframe(pd.DataFrame(cost_rows),use_container_width=True,hide_index=True)

    # ── Tab 3: Integrated Plan ────────────────────────────────────────────────
    with sc_tab3:
        st.markdown("#### 🌐 Multi-Project Integrated Cost Plan")
        st.caption(
            "Aggregates WBS-based cost distributions across selected subprojects onto a shared "
            "calendar axis. **Select subprojects only** (e.g. CM · CR · LP · PY) to avoid "
            "double-counting with any Full Proposal entry.")

        # ── Project selector ─────────────────────────────────────────────────
        all_projs_int = db_list()
        if not all_projs_int:
            st.warning("No projects in database."); st.stop()

        proj_display = {
            f"{p['project_name']}  [{p['doc']}  |  ฿{p['total_budget']:,.0f}]": p['key']
            for p in all_projs_int
        }
        # Default: all projects selected
        default_sel = list(proj_display.keys())

        sel_col, warn_col = st.columns([3, 1])
        selected_labels = sel_col.multiselect(
            "Projects to integrate:", options=list(proj_display.keys()),
            default=default_sel,
            help="Deselect 'Full Proposal' rows to prevent double-counting")

        if not selected_labels:
            st.warning("Select at least one project."); st.stop()

        selected_keys_int = [proj_display[lbl] for lbl in selected_labels]
        selected_data_int = [db_get(k) for k in selected_keys_int]
        selected_data_int = [d for d in selected_data_int if d]

        n_proj = len(selected_data_int)
        total_int_budget = sum(d.get("total_budget", 0) for d in selected_data_int)
        warn_col.markdown(
            f'<div class="kpi-card" style="margin-top:28px">'
            f'<div class="kpi-num" style="font-size:1.3rem;color:#0A0A0A">฿{total_int_budget:,.0f}</div>'
            f'<div class="kpi-label">Integrated Budget ({n_proj} projects)</div>'
            f'</div>', unsafe_allow_html=True)

        # ── Build global month axis ──────────────────────────────────────────
        proj_starts = [datetime.strptime(d["start_date"], "%Y-%m-%d") for d in selected_data_int]
        proj_ends   = [datetime.strptime(d["start_date"], "%Y-%m-%d") + relativedelta(months=d["n_months"]-1)
                       for d in selected_data_int]
        g_start = min(proj_starts)
        g_end   = max(proj_ends)
        N_int   = (g_end.year - g_start.year)*12 + (g_end.month - g_start.month) + 1
        int_labels = [(g_start + relativedelta(months=m)).strftime("%b %Y") for m in range(N_int)]

        # ── Aggregate across projects ────────────────────────────────────────
        int_mp   = [0.0] * N_int   # monthly planned cost (฿)
        int_ma   = [0.0] * N_int   # monthly actual cost (฿)
        int_ma_flag = [False] * N_int  # True if any project contributed actuals that month

        palette_proj = ["#0A0A0A","#1AE06B","#0070B8","#A06000","#CC2222",
                        "#6B6B6B","#00B4B4","#AA00AA","#E06000","#003080"]
        stacked_traces = []

        for idx, d in enumerate(selected_data_int):
            pname   = d.get("project_name", f"Project {idx+1}")
            p_start = datetime.strptime(d["start_date"], "%Y-%m-%d")
            offset  = (p_start.year - g_start.year)*12 + (p_start.month - g_start.month)
            budget  = d.get("total_budget", 0)
            acts_i  = d.get("activities", [])
            N_i     = d.get("n_months", 12)

            pm_proj = [0.0] * N_int   # this project's monthly planned cost
            for a in acts_i:
                s  = int(a["start_month"]); e = int(a["end_month"])
                dur = max(1, e - s + 1)
                # Activity cost = budget × weight%  (or use stored planned_cost)
                cost = a.get("planned_cost", 0) or round(a.get("weight", 0) / 100 * budget, 2)
                per_month = cost / dur
                for m in range(s, min(e + 1, N_i + 1)):
                    gi = offset + (m - 1)
                    if 0 <= gi < N_int:
                        int_mp[gi]   += per_month
                        pm_proj[gi]  += per_month
                # Actual costs
                for m_str, v in a.get("actual_costs", {}).items():
                    m_i = int(m_str)
                    gi  = offset + (m_i - 1)
                    if 0 <= gi < N_int:
                        int_ma[gi]      += float(v)
                        int_ma_flag[gi]  = True

            stacked_traces.append((pname, pm_proj, palette_proj[idx % len(palette_proj)]))

        # Cumulative arrays
        cum_int_p = []; cum_int_a = []
        rp = ra = 0.0
        for i in range(N_int):
            rp += int_mp[i]; cum_int_p.append(round(rp, 2))
            if int_ma_flag[i]:
                ra += int_ma[i]; cum_int_a.append(round(ra, 2))
            else:
                cum_int_a.append(round(ra, 2) if ra > 0 else None)

        burn_int = [total_int_budget - v for v in cum_int_p]

        total_int_actual  = next((v for v in reversed(cum_int_a) if v is not None), 0) or 0
        remaining_int     = total_int_budget - total_int_actual
        # Current month on global axis
        today = datetime.today()
        cm_int = max(1, min(
            (today.year - g_start.year)*12 + (today.month - g_start.month) + 1,
            N_int))

        pv_int = cum_int_p[cm_int - 1]
        cv_int = total_int_actual - pv_int

        # ── KPI row ──────────────────────────────────────────────────────────
        ik1, ik2, ik3, ik4, ik5 = st.columns(5)
        def _ik(col, val, lbl, ac="#0A0A0A"):
            col.markdown(f'<div class="kpi-card"><div class="kpi-num" style="color:{ac};font-size:1.5rem">'
                         f'{val}</div><div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)
        _ik(ik1, f"฿{total_int_budget:,.0f}",  "Integrated Budget")
        _ik(ik2, f"฿{pv_int:,.0f}",             f"Planned Cost (M now)")
        _ik(ik3, f"฿{total_int_actual:,.0f}",   "Actual Cost to Date",
            "#1AE06B" if total_int_actual <= pv_int else "#CC2222")
        _ik(ik4, f"฿{remaining_int:,.0f}",       "Budget Remaining",
            "#1AE06B" if remaining_int >= 0 else "#CC2222")
        _ik(ik5, f"฿{cv_int:+,.0f}",             "Cost Variance",
            "#1AE06B" if cv_int <= 0 else "#CC2222")

        st.markdown("")

        # ── Stacked bar (per-project monthly) + cumulative S-curve ──────────
        fig_int = go.Figure()

        # Stacked bars — monthly contribution per project
        for pname, pm_proj, col_p in stacked_traces:
            fig_int.add_trace(go.Bar(
                x=int_labels, y=pm_proj, name=pname,
                marker_color=col_p, opacity=0.85,
                hovertemplate="<b>%{x}</b><br>" + pname + ": ฿%{y:,.0f}<extra></extra>"))

        # Cumulative planned cost line (y2)
        fig_int.add_trace(go.Scatter(
            x=int_labels, y=cum_int_p, name="Cumulative Plan (฿)",
            line=dict(color="#0A0A0A", width=3),
            mode="lines+markers", marker=dict(size=6),
            yaxis="y2",
            hovertemplate="<b>%{x}</b><br>Cum Plan: ฿%{y:,.0f}<extra></extra>"))

        # Cumulative actual cost line (y2) — only where data exists
        ax_int = [int_labels[i] for i, v in enumerate(cum_int_a) if v is not None]
        ay_int = [v for v in cum_int_a if v is not None]
        if ay_int:
            fig_int.add_trace(go.Scatter(
                x=ax_int, y=ay_int, name="Cumulative Actual (฿)",
                line=dict(color="#1AE06B", width=3, dash="dot"),
                mode="lines+markers", marker=dict(size=8, symbol="diamond"),
                yaxis="y2",
                hovertemplate="<b>%{x}</b><br>Cum Actual: ฿%{y:,.0f}<extra></extra>"))

        # Burn-down line (y2)
        fig_int.add_trace(go.Scatter(
            x=int_labels, y=burn_int, name="Budget Remaining (฿)",
            line=dict(color="#CC2222", width=2, dash="dash"),
            mode="lines", yaxis="y2", visible="legendonly",
            hovertemplate="<b>%{x}</b><br>Remaining: ฿%{y:,.0f}<extra></extra>"))

        # Today vline
        if 1 <= cm_int <= N_int:
            fig_int.add_vline(x=int_labels[cm_int - 1],
                              line_dash="dash", line_color="#1AE06B", line_width=2)
            fig_int.add_annotation(x=int_labels[cm_int - 1], y=1, yref="paper",
                text="TODAY", showarrow=False,
                font=dict(color="#0A0A0A", size=11), xanchor="left")

        fig_int.update_layout(
            barmode="stack",
            paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A", size=12),
            height=450, margin=dict(l=20, r=20, t=20, b=20),
            hovermode="x unified",
            legend=dict(bgcolor="#FFFFFF", bordercolor="#E0E0E0", borderwidth=1,
                        orientation="h", yanchor="bottom", y=1.02, xanchor="left", x=0),
            xaxis=dict(gridcolor="#F0F0F0", showline=False, tickangle=-30),
            yaxis=dict(title="Monthly Cost (฿)", gridcolor="#F0F0F0",
                       tickprefix="฿", tickformat=",.0f"),
            yaxis2=dict(title="Cumulative Cost (฿)", overlaying="y", side="right",
                        tickprefix="฿", tickformat=",.0f",
                        gridcolor="#F0F0F0", showgrid=False))
        st.plotly_chart(fig_int, use_container_width=True)

        # ── Per-project summary table ────────────────────────────────────────
        st.markdown('<div class="sec">PROJECT CONTRIBUTION SUMMARY</div>', unsafe_allow_html=True)
        int_rows = []
        for d in selected_data_int:
            b = d.get("total_budget", 0)
            share = round(b / total_int_budget * 100, 1) if total_int_budget > 0 else 0
            n_acts_i = len(d.get("activities", []))
            wbs_cost_i = sum(
                a.get("planned_cost", 0) or round(a.get("weight", 0)/100*b, 2)
                for a in d.get("activities", []))
            actual_i = sum(
                float(v)
                for a in d.get("activities", [])
                for v in a.get("actual_costs", {}).values())
            int_rows.append({
                "Project": d.get("project_name", "?")[:40],
                "Doc": d.get("_db_doc","—"),
                "Budget (฿)": f"฿{b:,.0f}",
                "Share (%)": f"{share}%",
                "WBS Cost (฿)": f"฿{wbs_cost_i:,.0f}",
                "Actual to Date (฿)": f"฿{actual_i:,.0f}",
                "Remaining (฿)": f"฿{b - actual_i:,.0f}",
                "Activities": n_acts_i,
            })
        st.dataframe(pd.DataFrame(int_rows), use_container_width=True, hide_index=True)

        # ── Monthly detail table ─────────────────────────────────────────────
        with st.expander("📋 Monthly Integrated Cost Table"):
            detail_rows = []
            for i, lbl in enumerate(int_labels):
                detail_rows.append({
                    "Month": lbl,
                    "Monthly Plan (฿)": f"฿{int_mp[i]:,.0f}",
                    "Cumulative Plan (฿)": f"฿{cum_int_p[i]:,.0f}",
                    "Monthly Actual (฿)": f"฿{int_ma[i]:,.0f}" if int_ma_flag[i] else "—",
                    "Cumulative Actual (฿)": f"฿{cum_int_a[i]:,.0f}" if cum_int_a[i] else "—",
                    "Budget Remaining (฿)": f"฿{burn_int[i]:,.0f}",
                })
            st.dataframe(pd.DataFrame(detail_rows), use_container_width=True, hide_index=True)

# ── COST TRACKING ─────────────────────────────────────────────────────────────
elif page=="💰 Cost Tracking":
    st.markdown("# 💰 Cost Tracking")
    st.markdown("---")
    if not acts:
        st.info("Add activities first in ⚙️ Project Setup."); st.stop()

    budget_ct   = data.get("total_budget", 0)
    wbs_cost_ct = sum(a.get("planned_cost",0) or round(a.get("weight",0)/100*budget_ct,2) for a in acts)
    variance_ct = budget_ct - wbs_cost_ct
    pct_used_ct = round(wbs_cost_ct/budget_ct*100,1) if budget_ct>0 else 0

    # Actual cost to date from actual_costs entries
    total_actual_ct = next((v for v in reversed(cum_ac) if v is not None), 0) or 0
    pv_ct  = cum_pc[cm-1] if cm<=N else cum_pc[-1]
    cv_ct  = total_actual_ct - pv_ct
    remaining_ct = budget_ct - total_actual_ct

    # EVM cost metrics
    ev_pct_ct = (next((cuma[i] for i in range(cm-1,-1,-1) if cuma[i] is not None), 0) or 0)
    ev_cost_ct = budget_ct * ev_pct_ct / 100
    cpi_ct = round(ev_cost_ct / total_actual_ct, 2) if total_actual_ct > 0 else None

    # ── KPI row ───────────────────────────────────────────────────────────────
    k1,k2,k3,k4,k5,k6 = st.columns(6)
    def _ctk(col, val, lbl, ac="#0A0A0A"):
        col.markdown(f'<div class="kpi-card"><div class="kpi-num" style="color:{ac};font-size:1.45rem">'
                     f'{val}</div><div class="kpi-label">{lbl}</div></div>', unsafe_allow_html=True)
    _ctk(k1, f"฿{budget_ct:,.0f}",        "Contract Budget")
    _ctk(k2, f"฿{wbs_cost_ct:,.0f}",      "WBS Planned Cost",
         "#1AE06B" if wbs_cost_ct<=budget_ct else "#CC2222")
    _ctk(k3, f"฿{pv_ct:,.0f}",            f"Planned Cost M{cm}")
    _ctk(k4, f"฿{total_actual_ct:,.0f}",  f"Actual Cost M{cm}",
         "#1AE06B" if total_actual_ct<=pv_ct else "#CC2222")
    _ctk(k5, f"฿{remaining_ct:,.0f}",     "Budget Remaining",
         "#1AE06B" if remaining_ct>=0 else "#CC2222")
    _ctk(k6, f"{cpi_ct:.2f}" if cpi_ct else "—", "CPI (EV/AC)",
         "#1AE06B" if cpi_ct and cpi_ct>=1 else "#A06000" if cpi_ct and cpi_ct>=0.8 else "#CC2222")

    st.markdown("")

    # ── Budget utilisation bar ────────────────────────────────────────────────
    bar_col_ct = "#1AE06B" if pct_used_ct<=100 else "#CC2222"
    st.markdown(
        f'<div style="background:#FFFFFF;border-radius:16px;padding:16px 20px;'
        f'box-shadow:0 2px 8px rgba(0,0,0,.06);margin-bottom:16px">'
        f'<div style="display:flex;justify-content:space-between;margin-bottom:8px">'
        f'<span style="font-weight:700;color:#0A0A0A">WBS Cost vs Contract Budget</span>'
        f'<span style="font-weight:800;font-size:1.1rem;color:{bar_col_ct}">{pct_used_ct}%</span></div>'
        f'<div style="background:#EFEFEF;border-radius:8px;height:10px">'
        f'<div style="background:{bar_col_ct};height:10px;border-radius:8px;width:{min(pct_used_ct,100):.1f}%"></div></div>'
        f'<div style="display:flex;justify-content:space-between;margin-top:6px">'
        f'<span style="font-size:.75rem;color:#6B6B6B">฿0</span>'
        f'<span style="font-size:.75rem;color:#6B6B6B">฿{budget_ct:,.0f}</span></div>'
        f'</div>', unsafe_allow_html=True)

    # ── Cost S-curve + monthly bar ────────────────────────────────────────────
    ct_tab1, ct_tab2, ct_tab3 = st.tabs(["📈 Cost S-Curve", "📊 Monthly Breakdown", "📋 Activity Detail"])

    with ct_tab1:
        fig_cs = go.Figure()
        fig_cs.add_trace(go.Scatter(
            x=labels, y=cum_pc, name="Planned Cumulative (฿)",
            fill="tozeroy", fillcolor="rgba(10,10,10,.07)",
            line=dict(color="#0A0A0A", width=2.5), mode="lines+markers", marker=dict(size=7),
            hovertemplate="<b>%{x}</b><br>Plan: ฿%{y:,.0f}<extra></extra>"))
        ax_ac=[labels[i] for i,v in enumerate(cum_ac) if v is not None]
        ay_ac=[v for v in cum_ac if v is not None]
        if ay_ac:
            fig_cs.add_trace(go.Scatter(
                x=ax_ac, y=ay_ac, name="Actual Cumulative (฿)",
                fill="tozeroy", fillcolor="rgba(26,224,107,.15)",
                line=dict(color="#1AE06B", width=2.5), mode="lines+markers",
                marker=dict(size=9, symbol="diamond"),
                hovertemplate="<b>%{x}</b><br>Actual: ฿%{y:,.0f}<extra></extra>"))
        # Burn-down
        burn_ct = [budget_ct - v for v in cum_pc]
        fig_cs.add_trace(go.Scatter(
            x=labels, y=burn_ct, name="Budget Remaining (฿)",
            line=dict(color="#CC2222", width=1.8, dash="dash"),
            mode="lines", yaxis="y2", visible="legendonly",
            hovertemplate="<b>%{x}</b><br>Remaining: ฿%{y:,.0f}<extra></extra>"))
        if cm<=N:
            fig_cs.add_vline(x=labels[cm-1], line_dash="dash", line_color="#1AE06B", line_width=2)
            fig_cs.add_annotation(x=labels[cm-1], y=1, yref="paper", text="TODAY",
                showarrow=False, font=dict(color="#0A0A0A", size=11), xanchor="left")
        fig_cs.update_layout(
            paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A", size=12), height=400,
            margin=dict(l=20,r=20,t=20,b=20), hovermode="x unified",
            legend=dict(bgcolor="#FFFFFF", bordercolor="#E0E0E0", borderwidth=1),
            xaxis=dict(gridcolor="#F0F0F0", showline=False, tickangle=-30),
            yaxis=dict(title="Cumulative Cost (฿)", gridcolor="#F0F0F0", tickprefix="฿", tickformat=",.0f"),
            yaxis2=dict(title="Remaining (฿)", overlaying="y", side="right",
                        tickprefix="฿", tickformat=",.0f", showgrid=False))
        st.plotly_chart(fig_cs, use_container_width=True)

    with ct_tab2:
        fig_mb = go.Figure()
        fig_mb.add_trace(go.Bar(
            x=labels, y=monthly_pc, name="Planned Spend (฿)",
            marker_color="#0A0A0A", opacity=0.8,
            hovertemplate="<b>%{x}</b><br>Plan: ฿%{y:,.0f}<extra></extra>"))
        # actual monthly
        ma_bar = [monthly_ac[i] if monthly_ac[i] is not None else 0 for i in range(N)]
        if any(v and v>0 for v in ma_bar):
            fig_mb.add_trace(go.Bar(
                x=labels, y=ma_bar, name="Actual Spend (฿)",
                marker_color="#1AE06B", opacity=0.85,
                hovertemplate="<b>%{x}</b><br>Actual: ฿%{y:,.0f}<extra></extra>"))
        if cm<=N:
            fig_mb.add_vline(x=labels[cm-1], line_dash="dash", line_color="#1AE06B", line_width=1.5)
        fig_mb.update_layout(
            barmode="group", paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A", size=12), height=380,
            margin=dict(l=20,r=20,t=20,b=20), hovermode="x unified",
            legend=dict(bgcolor="#FFFFFF", bordercolor="#E0E0E0", borderwidth=1),
            xaxis=dict(gridcolor="#F0F0F0", showline=False, tickangle=-30),
            yaxis=dict(title="Monthly Spend (฿)", gridcolor="#F0F0F0", tickprefix="฿", tickformat=",.0f"))
        st.plotly_chart(fig_mb, use_container_width=True)

        # Monthly table
        mt_rows = []
        for i, lbl in enumerate(labels):
            ac_v = monthly_ac[i] or 0
            cv_m = ac_v - monthly_pc[i]
            mt_rows.append({"Month": lbl,
                "Plan (฿)": f"฿{monthly_pc[i]:,.0f}",
                "Actual (฿)": f"฿{ac_v:,.0f}" if ac_v else "—",
                "Variance (฿)": f"฿{cv_m:+,.0f}" if ac_v else "—",
                "Cum Plan (฿)": f"฿{cum_pc[i]:,.0f}",
                "Cum Actual (฿)": f"฿{cum_ac[i]:,.0f}" if cum_ac[i] else "—"})
        st.dataframe(pd.DataFrame(mt_rows), use_container_width=True, hide_index=True)

    with ct_tab3:
        act_rows_ct = []
        for a in acts:
            pc_a = a.get("planned_cost",0) or round(a.get("weight",0)/100*budget_ct,2)
            ac_a = sum(float(v) for v in a.get("actual_costs",{}).values())
            share = round(pc_a/wbs_cost_ct*100,1) if wbs_cost_ct>0 else 0
            rem_a = pc_a - ac_a
            cpi_a = round(pc_a/ac_a,2) if ac_a>0 else None
            act_rows_ct.append({
                "No": a["no"], "Activity": a["name"][:45],
                "Wt%": a["weight"],
                "Budget (฿)": f"฿{pc_a:,.0f}",
                "Actual (฿)": f"฿{ac_a:,.0f}",
                "Remaining (฿)": f"฿{rem_a:,.0f}",
                "Share%": f"{share}%",
                "CPI": f"{cpi_a:.2f}" if cpi_a else "—",
                "Status": a.get("status","Not Started")
            })
        st.dataframe(pd.DataFrame(act_rows_ct), use_container_width=True, hide_index=True)

        # Horizontal bar — budget vs actual per activity
        act_names_ct = [f"{a['no']}. {a['name'][:30]}" for a in acts]
        budgets_ct   = [a.get("planned_cost",0) or round(a.get("weight",0)/100*budget_ct,2) for a in acts]
        actuals_ct   = [sum(float(v) for v in a.get("actual_costs",{}).values()) for a in acts]
        fig_act = go.Figure()
        fig_act.add_trace(go.Bar(y=act_names_ct, x=budgets_ct, name="Budget",
            orientation="h", marker_color="#0A0A0A", opacity=0.7,
            hovertemplate="<b>%{y}</b><br>Budget: ฿%{x:,.0f}<extra></extra>"))
        if any(v>0 for v in actuals_ct):
            fig_act.add_trace(go.Bar(y=act_names_ct, x=actuals_ct, name="Actual",
                orientation="h", marker_color="#1AE06B",
                hovertemplate="<b>%{y}</b><br>Actual: ฿%{x:,.0f}<extra></extra>"))
        fig_act.update_layout(
            barmode="overlay", paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A", size=11),
            height=max(300, len(acts)*38),
            margin=dict(l=10,r=80,t=10,b=10),
            xaxis=dict(gridcolor="#F0F0F0", showline=False, tickprefix="฿", tickformat=",.0f"),
            yaxis=dict(autorange="reversed"),
            legend=dict(bgcolor="#FFFFFF", bordercolor="#E0E0E0", borderwidth=1))
        st.plotly_chart(fig_act, use_container_width=True)

# ── GANTT ─────────────────────────────────────────────────────────────────────
elif page=="📅 Gantt":
    st.markdown("# 📅 Gantt Chart"); st.markdown("---")
    base=datetime.strptime(SI,"%Y-%m-%d"); rows=[]
    for a in acts:
        s=base+relativedelta(months=a["start_month"]-1,day=1)
        e=base+relativedelta(months=a["end_month"],day=1)-relativedelta(days=1)
        sk=next((k for k in SCOL if k in a.get("status","")),"Not Started")
        rows.append(dict(Task=f"{a['no']}. {a['name'][:45]}",Start=s.strftime("%Y-%m-%d"),
                         Finish=e.strftime("%Y-%m-%d"),Status=sk,Weight=f"{a['weight']}%"))
    if rows:
        df=pd.DataFrame(rows)
        fig=px.timeline(df,x_start="Start",x_end="Finish",y="Task",color="Status",
                        color_discrete_map=SCOL,hover_data=["Weight"])
        fig.update_yaxes(autorange="reversed")
        fig.add_vline(x=date.today().isoformat(),line_dash="dash",line_color="#1AE06B",line_width=2)
        fig.add_annotation(x=date.today().isoformat(),y=1.02,yref="paper",
            text="TODAY",showarrow=False,font=dict(color="#0A0A0A",size=11),xanchor="left")
        fig.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",
            font=dict(color="#0A0A0A",size=11),height=420,margin=dict(l=10,r=10,t=20,b=10),
            xaxis=dict(gridcolor="#F0F0F0",showline=False),yaxis=dict(gridcolor="#F0F0F0"),legend=dict(bgcolor="#FFFFFF"))
        st.plotly_chart(fig,use_container_width=True)
    else:
        st.info("No activities yet. Add them in ⚙️ Project Setup.")

# ── EVM ───────────────────────────────────────────────────────────────────────
elif page=="📊 EVM Indicators":
    st.markdown("# 📊 EVM Indicators"); st.markdown("---")

    # Cost summary from WBS
    _total_budget = data.get("total_budget", 0)
    _wbs_cost = sum(a.get("planned_cost", 0) for a in acts)
    _ev_pct = cuma[cm-1] if cm <= N and cuma[cm-1] is not None else 0
    _pv_pct = cump[cm-1] if cm <= N else 0
    # EV cost = budget × EV%/100
    _ev_cost = _total_budget * _ev_pct / 100
    _pv_cost = _total_budget * _pv_pct / 100
    _cv = _ev_cost - _pv_cost  # Cost Variance (simplified: no actual cost input)

    ec1, ec2, ec3, ec4 = st.columns(4)
    ec1.metric("Project Budget", f"฿{_total_budget:,.0f}")
    ec2.metric("WBS Planned Cost", f"฿{_wbs_cost:,.0f}",
               delta=f"฿{_wbs_cost - _total_budget:+,.0f}" if _wbs_cost else None, delta_color="inverse")
    ec3.metric(f"PV Cost (M{cm})", f"฿{_pv_cost:,.0f}")
    ec4.metric(f"EV Cost (M{cm})", f"฿{_ev_cost:,.0f}", delta=f"฿{_cv:+,.0f}", delta_color="normal")

    st.markdown("---")
    st.markdown("|Indicator|Formula|Meaning|\n|---|---|---|\n"
        "|**PV**|Cumulative Plan %|What should be done|\n"
        "|**EV**|Cumulative Actual %|What has been done|\n"
        "|**SV**|EV − PV|+ ahead / − delayed|\n"
        "|**SPI**|EV / PV|>1 ahead • <1 behind|\n"
        "|**CV**|EV cost − PV cost|Budget performance|\n"
        "|**WBS Cost**|Σ Activity planned cost|Bottom-up budget check|")
    st.markdown("---")
    rows=[]
    for i,lbl in enumerate(labels):
        pv=cump[i]; ev=cuma[i]
        if ev is None: rows.append({"Month":lbl,"PV":f"{pv:.1f}","EV":"—","SV":"—","SPI":"—","Status":"—"})
        else:
            sv_=round(ev-pv,2); spi_=round(ev/pv,2) if pv>0 else None
            rows.append({"Month":lbl,"PV":f"{pv:.1f}","EV":f"{ev:.1f}","SV":f"{sv_:+.1f}",
                "SPI":f"{spi_:.2f}" if spi_ else "—",
                "Status":("✅ On Track" if spi_ and spi_>=1.05 else "⚠️ Warning" if spi_ and spi_>=0.95
                          else "💤 Delayed" if spi_ and spi_>=0.8 else "🔴 Critical" if spi_ else "—")})
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    spi_v=[round(cuma[i]/cump[i],2) if (cuma[i] is not None and cump[i]>0) else None for i in range(N)]
    xs=[labels[i] for i,v in enumerate(spi_v) if v is not None]; ys=[v for v in spi_v if v is not None]
    if ys:
        fig=go.Figure()
        fig.add_hline(y=1.0,line_color="#1AE06B",line_width=2)
        fig.add_annotation(x=1,xref="paper",y=1.0,text="Target SPI=1.0",showarrow=False,
            font=dict(color="#0A0A0A",size=10),xanchor="right",yanchor="bottom")
        fig.add_hline(y=0.95,line_dash="dot",line_color="#A06000")
        fig.add_annotation(x=1,xref="paper",y=0.95,text="⚠️ Warning",showarrow=False,
            font=dict(color="#A06000",size=10),xanchor="right",yanchor="bottom")
        fig.add_hline(y=0.80,line_dash="dot",line_color="#CC2222")
        fig.add_annotation(x=1,xref="paper",y=0.80,text="🔴 Critical",showarrow=False,
            font=dict(color="#CC2222",size=10),xanchor="right",yanchor="bottom")
        fig.add_trace(go.Scatter(x=xs,y=ys,name="SPI",line=dict(color="#0A0A0A",width=2.5),
            mode="lines+markers",marker=dict(size=8,color=["#1AE06B" if v>=1 else "#A06000" if v>=0.8 else "#CC2222" for v in ys])))
        fig.update_layout(paper_bgcolor="#FFFFFF",plot_bgcolor="#FFFFFF",font=dict(color="#0A0A0A"),
            height=260,margin=dict(l=10,r=10,t=10,b=10),showlegend=False,
            xaxis=dict(gridcolor="#F0F0F0",showline=False),yaxis=dict(gridcolor="#F0F0F0",title="SPI"))
        st.plotly_chart(fig,use_container_width=True)


# ── UPDATE PROGRESS ───────────────────────────────────────────────────────────
elif page=="📝 Update Progress":
    st.markdown("# 📝 Update Progress")
    st.markdown("---")
    if not acts: st.info("Add activities first in ⚙️ Project Setup."); st.stop()

    sopts_raw = ["Not Started","In Progress","Pending","Completed","Delayed","Cancelled"]

    # ── Quick Activity Editor ─────────────────────────────────────────────────
    with st.expander("✏️ Edit Activities (Name / Weight / Dates / Cost)", expanded=True):
        st.caption("Modify activity details here without going to Project Setup. Click 💾 Save Activities when done.")
        sopts_up = ["Not Started","In Progress","Pending","Completed","Delayed","Cancelled"]
        up_rows = []
        for a in acts:
            sk2 = next((k for k in sopts_up if k in a.get("status","")), sopts_up[0])
            up_rows.append({
                "No":           a.get("no",""),
                "Name (EN)":    a.get("name",""),
                "Weight %":     float(a.get("weight", 0)),
                "Planned Cost": float(a.get("planned_cost", 0)),
                "Start M":      int(a.get("start_month", 1)),
                "End M":        int(a.get("end_month", 1)),
                "Status":       sk2,
            })
        up_df = pd.DataFrame(up_rows) if up_rows else pd.DataFrame(
            columns=["No","Name (EN)","Weight %","Planned Cost","Start M","End M","Status"])

        up_edited = st.data_editor(
            up_df, use_container_width=True, hide_index=True, num_rows="fixed",
            column_config={
                "No":           st.column_config.TextColumn("No",        width="small",  disabled=True),
                "Name (EN)":    st.column_config.TextColumn("Name (EN)", width="large"),
                "Weight %":     st.column_config.NumberColumn("Weight %", min_value=0.0, max_value=100.0, step=0.5, format="%.2f", width="small"),
                "Planned Cost": st.column_config.NumberColumn("Budget (฿)", min_value=0.0, step=10000.0, format="฿%.0f", width="medium"),
                "Start M":      st.column_config.NumberColumn("Start M", min_value=1, max_value=36, step=1, width="small"),
                "End M":        st.column_config.NumberColumn("End M",   min_value=1, max_value=36, step=1, width="small"),
                "Status":       st.column_config.SelectboxColumn("Status", width="medium", options=sopts_up),
            },
            key="up_act_editor"
        )

        wt_sum_up = up_edited["Weight %"].sum() if len(up_edited) > 0 else 0
        ua_info, ua_norm, ua_save = st.columns([3, 1, 1])
        ua_info.markdown(
            f"**{len(up_edited)} activities** · Weight total: **{wt_sum_up:.1f}%** "
            f"{'✅' if abs(wt_sum_up - 100) < 1 else '⚠️ should sum to 100%'}")
        if ua_norm.button("⚖️ Normalize", use_container_width=True, key="ua_norm"):
            if wt_sum_up > 0:
                up_edited["Weight %"] = (up_edited["Weight %"] / wt_sum_up * 100).round(2)
                st.info("Normalized. Click 💾 Save Activities to apply.")
        if ua_save.button("💾 Save Activities", use_container_width=True, type="primary", key="ua_save"):
            for i, row in up_edited.iterrows():
                name_v = str(row.get("Name (EN)","")).strip()
                if not name_v: continue
                st_raw2 = str(row.get("Status","Not Started"))
                data["activities"][i]["name"]         = name_v
                data["activities"][i]["weight"]        = float(row.get("Weight %", 0) or 0)
                data["activities"][i]["planned_cost"]  = float(row.get("Planned Cost", 0) or 0)
                data["activities"][i]["start_month"]   = int(row.get("Start M", 1) or 1)
                data["activities"][i]["end_month"]     = int(row.get("End M", 1) or 1)
                data["activities"][i]["status"]        = SMAP.get(st_raw2, st_raw2)
            save_data(data)
            st.success("✅ Activities saved!"); st.rerun()

    st.markdown(f"Double-click any cell to edit. Current month: **M{cm} — {labels[cm-1] if cm<=N else 'N/A'}** (highlighted).")

    # Build single wide table: rows = activities, cols = Status + all months
    prows = []
    for a in acts:
        sk = next((k for k in sopts_raw if k in a.get("status","")), sopts_raw[0])
        row = {"No": a["no"], "Activity": a["name"][:45], "Wt%": a["weight"],
               "Cost(฿)": a.get("planned_cost", 0.0), "Status": sk}
        for m in range(1, N+1):
            in_range = a["start_month"] <= m <= a["end_month"]
            row[f"M{m}"] = float(a.get("actuals",{}).get(str(m), 0.0)) if in_range else None
        prows.append(row)

    pdf = pd.DataFrame(prows)

    # Month column configs — highlight current month
    m_cfg = {}
    for m in range(1, N+1):
        col_lbl = f"M{m} ★" if m == cm else f"M{m}"
        m_cfg[f"M{m}"] = st.column_config.NumberColumn(
            col_lbl, min_value=0.0, max_value=100.0,
            step=5.0, format="%.0f", width="small")

    col_cfg = {
        "No":       st.column_config.TextColumn("No",         width="small",  disabled=True),
        "Activity": st.column_config.TextColumn("Activity",   width="large",  disabled=True),
        "Wt%":      st.column_config.NumberColumn("Wt%",      width="small",  disabled=True, format="%.1f"),
        "Cost(฿)":  st.column_config.NumberColumn("Cost(฿)",  width="medium", disabled=True, format="฿%.0f"),
        "Status":   st.column_config.SelectboxColumn("Status", width="medium", options=sopts_raw),
        **m_cfg,
    }

    edited = st.data_editor(
        pdf,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config=col_cfg,
        key="progress_editor"
    )

    st.markdown("")
    cl, cr = st.columns([4, 1])
    cl.info("💡 Edit **Status** and any **M1–M12** cell. Blank cells (grey) = outside activity date range — leave as blank.")
    if cr.button("💾 Save All", use_container_width=True, type="primary"):
        for i, row in edited.iterrows():
            st_raw = str(row["Status"])
            data["activities"][i]["status"] = SMAP.get(st_raw, st_raw)
            for m in range(1, N+1):
                val = row.get(f"M{m}")
                if val is not None and not pd.isna(val):
                    data["activities"][i].setdefault("actuals",{})[str(m)] = float(val)
        save_data(data)
        st.success("✅ Progress saved!"); st.rerun()

    # ── Actual Cost Input ──────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 💰 Actual Cost per Month (฿)")
    budget_up = data.get("total_budget", 0)

    wbh1, wbh2 = st.columns([5,1])
    wbh1.caption("Enter actual expenditure (฿) per month per activity. Leave 0 if not yet spent.")
    if wbh2.button("📐 Fill from Weight", help="Auto-set Budget: Weight% × Total Budget",
                   use_container_width=True):
        for a in data["activities"]:
            a["planned_cost"] = round(a.get("weight",0)/100*budget_up, 2)
        save_data(data); st.success("✅ Budgets computed from weights."); st.rerun()

    crows = []
    for a in acts:
        pc = a.get("planned_cost",0) or round(a.get("weight",0)/100*budget_up, 2)
        row = {"No": a["no"], "Activity": a["name"][:40], "Budget": float(pc)}
        for m in range(1, N+1):
            in_range = a["start_month"] <= m <= a["end_month"]
            row[f"C{m}"] = float(a.get("actual_costs",{}).get(str(m), 0.0)) if in_range else None
        crows.append(row)

    cdf = pd.DataFrame(crows)
    # Force numeric columns to float64 so st.data_editor renders them as editable
    cdf["Budget"] = pd.to_numeric(cdf["Budget"], errors="coerce").fillna(0.0)
    for m in range(1, N+1):
        cdf[f"C{m}"] = pd.to_numeric(cdf[f"C{m}"], errors="coerce")

    # Determine which months have at least one activity in range
    months_in_range = set()
    for a in acts:
        for m in range(a.get("start_month", 1), a.get("end_month", 1) + 1):
            months_in_range.add(m)
    # Month columns with NO activity → fully disable that column
    locked_month_cols = [f"C{m}" for m in range(1, N+1) if m not in months_in_range]
    all_disabled_cols  = ["No", "Activity"] + locked_month_cols

    cm_cfg2 = {}
    for m in range(1, N+1):
        if m not in months_in_range:
            lbl2 = "🔒"          # column is fully locked — no activity scheduled
        elif m == cm:
            lbl2 = f"C{m} ★"
        else:
            lbl2 = f"C{m}"
        cm_cfg2[f"C{m}"] = st.column_config.NumberColumn(
            lbl2, min_value=0.0, step=1000.0, format="%.0f", width="small")
    cost_cfg = {
        "No":       st.column_config.TextColumn("No",       width="small",  disabled=True),
        "Activity": st.column_config.TextColumn("Activity", width="large",  disabled=True),
        "Budget":   st.column_config.NumberColumn("Budget (฿) ✏️", width="medium", min_value=0.0, step=1000.0, format="%.0f"),
        **cm_cfg2,
    }
    cost_edited = st.data_editor(cdf, use_container_width=True, hide_index=True,
                                  num_rows="fixed", column_config=cost_cfg,
                                  disabled=all_disabled_cols, key="cost_editor")
    cl2, cr2 = st.columns([4,1])
    cl2.caption("★ = current month. Blank cells (outside activity schedule) are locked and won't be saved.")
    if cr2.button("💾 Save Costs", use_container_width=True, type="primary"):
        for i, row in cost_edited.iterrows():
            # Save edited planned_cost (Budget column)
            budget_val = row.get("Budget")
            if budget_val is not None and not pd.isna(budget_val):
                data["activities"][i]["planned_cost"] = float(budget_val)
            # Save actual costs per month
            a_start = data["activities"][i].get("start_month", 1)
            a_end   = data["activities"][i].get("end_month", N)
            for m in range(1, N+1):
                val = row.get(f"C{m}")
                # Only save values within the activity's scheduled range
                if val is not None and not pd.isna(val) and a_start <= m <= a_end:
                    data["activities"][i].setdefault("actual_costs",{})[str(m)] = float(val)
        save_data(data); st.success("✅ Costs saved!"); st.rerun()

    # ── Excel Export / Import ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📊 Excel Export / Import")
    ex_col, im_col = st.columns(2)

    with ex_col:
        st.markdown("**📥 Export to Excel**")
        fname = re.sub(r'[^A-Za-z0-9_]','_', data.get("project_name","project"))[:30]
        fname = f"{fname}_progress.xlsx"
        xl_bytes = build_excel(data, cump, cuma, labels)
        st.download_button(
            label="⬇️ Download Excel",
            data=xl_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        st.caption("3 sheets: **Progress** (editable), **S-Curve Data**, **Project Info**. Edit Status & month cells, then import back below.")

    with im_col:
        st.markdown("**📤 Import from Excel**")
        xl_up = st.file_uploader("Upload edited Excel", type=["xlsx"], key="xl_import")
        if xl_up:
            try:
                import openpyxl as _xl
                _wb  = _xl.load_workbook(xl_up, data_only=True)
                _ws  = _wb["Progress"]
                _rows = list(_ws.iter_rows(min_row=2, values_only=True))
                sopts_xl = ["Not Started","In Progress","Pending","Completed","Delayed","Cancelled"]
                updated = 0
                for ri, xrow in enumerate(_rows):
                    if ri >= len(data["activities"]): break
                    # col 5 = Status (0-indexed: col 4)
                    st_xl = str(xrow[4]).strip() if xrow[4] else "Not Started"
                    st_xl = next((k for k in sopts_xl if k.lower() in st_xl.lower()), "Not Started")
                    data["activities"][ri]["status"] = SMAP.get(st_xl, st_xl)
                    # cols 6..N+5 = M1..MN (0-indexed: 5..N+4)
                    for mi in range(1, N+1):
                        val = xrow[4+mi] if (4+mi) < len(xrow) else None
                        if val is not None and str(val).strip() not in ("","None"):
                            try:
                                fval = float(val)
                                if 0 <= fval <= 100:
                                    data["activities"][ri].setdefault("actuals",{})[str(mi)] = fval
                            except (ValueError, TypeError):
                                pass
                    updated += 1
                save_data(data)
                st.success(f"✅ Imported {updated} rows from Excel and saved!")
                st.rerun()
            except Exception as _e:
                st.error(f"❌ Could not read Excel: {_e}")

# ── WBS LIBRARY ───────────────────────────────────────────────────────────────
elif page=="📚 WBS Library":
    st.markdown("# 📚 WBS Library")
    st.markdown("---")

    MANIFEST_PATH_LIB = "projects/project_manifest.json"
    PROJECTS_DIR_LIB  = "projects"

    # ── Search + Sort controls ────────────────────────────────────────────────
    hc1, hc2, hc3 = st.columns([3, 2, 1])
    search_q  = hc1.text_input("🔍 Search", placeholder="project name, doc code…", label_visibility="collapsed")
    sort_by   = hc2.selectbox("Sort by", ["updated_at","created_at","project_name","n_activities"],
                               format_func=lambda x: {"updated_at":"Last Modified","created_at":"Date Added",
                                                       "project_name":"Name","n_activities":"No. Tasks"}[x],
                               label_visibility="collapsed")
    doc_icons = {"FULL":"🌐","CM":"🏔️","CR":"🌾","LP":"🌿","PY":"💧"}

    entries = db_list(sort=sort_by, search=search_q)
    active_key = data.get("_db_key","")

    if not entries:
        st.info("No projects in library yet. Use **➕ Add Current Project** below or run `wbs_extractor.py`.")
    else:
        hc3.markdown(f"<div style='text-align:right;padding-top:8px;color:#6B7280'>{len(entries)} projects</div>",
                     unsafe_allow_html=True)
        active_entry = next((e for e in entries if e["key"] == active_key), None)
        if active_entry:
            st.markdown(
                f"<div style='background:#0A0A0A;border-radius:12px;padding:10px 16px;margin-bottom:10px'>"
                f"<span style='color:#1AE06B;font-weight:700'>▶ ACTIVE</span>&nbsp;&nbsp;"
                f"<span style='color:#FFFFFF;font-weight:600'>{active_entry['project_name']}</span>"
                f"<span style='color:#9B9B9B;font-size:.8rem'> ({active_entry['doc']} / {active_entry['phase']} · "
                f"{active_entry['n_activities']} tasks)</span></div>", unsafe_allow_html=True)
        st.markdown("---")

        for e in entries:
            icon      = doc_icons.get(e["doc"],"📁")
            upd       = e["updated_at"][:16] if e["updated_at"] else "—"
            cre       = e["created_at"][:16] if e["created_at"] else "—"
            is_active = e["key"] == active_key
            row_bg    = "background:#FFFFFF;border:1.5px solid #1AE06B;border-radius:12px;padding:4px 8px;margin-bottom:2px" if is_active else ""
            active_badge = "&nbsp;<span style='background:#0A0A0A;color:#1AE06B;font-size:.68rem;font-weight:700;padding:2px 8px;border-radius:10px'>● IN USE</span>" if is_active else ""
            if row_bg:
                st.markdown(f"<div style='{row_bg}'>", unsafe_allow_html=True)
            c_info, c_acts, c_bud, c_upd, c_load, c_del = st.columns([4, 1, 2, 2, 1, 1])
            c_info.markdown(f"{icon} **{e['project_name']}**{active_badge}  \n<small style='color:#6B7280'>{e['doc']} / {e['phase']}</small>", unsafe_allow_html=True)
            c_acts.metric("Tasks", e["n_activities"])
            c_bud.metric("Budget", "฿{:,.0f}".format(e["total_budget"]) if e["total_budget"] else "—")
            c_upd.markdown(f"<div style='font-size:.78rem;color:#6B7280;padding-top:6px'>✏️ {upd}<br>➕ {cre}</div>", unsafe_allow_html=True)
            if row_bg:
                st.markdown("</div>", unsafe_allow_html=True)
            if c_load.button("✅ Active" if is_active else "▶ Load", key=f"lib_load_{e['key']}",
                              disabled=is_active, type="primary" if is_active else "secondary"):
                proj = db_get(e["key"])
                if proj:
                    proj["_db_key"] = e["key"]; proj["_db_doc"] = e["doc"]; proj["_db_phase"] = e["phase"]
                    save_data(proj); db_touch(e["key"])
                    st.success(f"✅ Loaded **{e['project_name']}** ({e['n_activities']} tasks)"); st.rerun()
                else:
                    st.error("Project data not found.")
            if c_del.button("🗑️", key=f"lib_del_{e['key']}", help=f"Delete '{e['project_name']}'"):
                db_log(e["key"], e["project_name"], "🗑️ Deleted", "Removed from library")
                db_delete(e["key"]); st.success(f"🗑️ Deleted **{e['project_name']}**"); st.rerun()
            st.markdown("<hr style='margin:6px 0;border-color:#E5E7EB'>", unsafe_allow_html=True)

    # ── Bulk Import ───────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📦 Import All WBS Extract Files")
    bulk_tab1, bulk_tab2 = st.tabs(["📁 From projects/ folder (JSON)", "📄 From CSV Template"])

    with bulk_tab1:
        st.markdown("Scan the `projects/` folder and add every JSON file into the database in one click.")
        imp_col1, imp_col2 = st.columns([3, 1])
        json_files_lib = [f for f in os.listdir(PROJECTS_DIR_LIB) if f.endswith(".json") and f != "project_manifest.json"] if os.path.exists(PROJECTS_DIR_LIB) else []
        imp_col1.markdown(f"<div style='padding:8px 0;color:#6B7280;font-size:.88rem'>Found <b>{len(json_files_lib)}</b> JSON files · <b>{len(db_list())}</b> in database</div>", unsafe_allow_html=True)
        if imp_col2.button("📦 Import All JSON", use_container_width=True, type="primary", disabled=(len(json_files_lib)==0), key="lib_bulk_json"):
            imported = errors = 0
            existing_keys = {e["key"] for e in db_list()}
            manifest_map = {}
            if os.path.exists(MANIFEST_PATH_LIB):
                with open(MANIFEST_PATH_LIB, encoding="utf-8") as _mf:
                    for m in json.load(_mf): manifest_map[m.get("filename","")] = m
            for fname in json_files_lib:
                try:
                    with open(os.path.join(PROJECTS_DIR_LIB, fname), encoding="utf-8") as _pf:
                        proj = json.load(_pf)
                    meta    = manifest_map.get(fname, {})
                    raw_key = meta.get("key") or re.sub(r'[^A-Za-z0-9_]','_', fname.replace(".json",""))[:40]
                    pname   = meta.get("project_name") or proj.get("project_name", fname)
                    pnameth = meta.get("project_name_th","") or proj.get("project_name_th","")
                    doc     = meta.get("doc") or proj.get("_db_doc","CUSTOM")
                    phase   = meta.get("phase") or proj.get("_db_phase","Custom")
                    proj["_db_key"] = raw_key; proj["_db_doc"] = doc; proj["_db_phase"] = phase
                    db_save(raw_key, pname, pnameth, doc, phase, proj)
                    db_log(raw_key, pname, "📦 Bulk Import", fname); imported += 1
                except Exception: errors += 1
            st.success(f"✅ Imported **{imported}** projects | ⚠️ Errors: {errors}"); st.rerun()

    with bulk_tab2:
        st.markdown("Upload `heymorning_task_import.csv` — every **PHASE** becomes a separate project.")
        def _ext(notes, key):
            m = re.search(rf'{key}: ([^\|]+)', notes or ""); return m.group(1).strip() if m else ""
        def _mnths(notes):
            raw = _ext(notes, "Months")
            try: return [int(x) for x in raw.split(",") if x.strip()]
            except: return []
        def _wt(notes):
            try: return float(_ext(notes, "Weight %"))
            except: return 0.0
        def csv_to_proj(rows, proj_name, start_date, n_months, budget):
            acts = []
            for i, r in enumerate(rows):
                mnths = _mnths(r.get("NOTES",""))
                if not mnths: continue
                wt = _wt(r.get("NOTES",""))
                wbs = _ext(r.get("NOTES",""), "WBS")
                try: pc = float(_ext(r.get("NOTES",""), "Estimated cost"))
                except: pc = 0.0
                acts.append({"no": wbs or str(i+1), "name": r.get("TASK NAME","").strip(),
                              "name_th": r.get("TASK NAME","").strip(), "weight": wt, "planned_cost": pc,
                              "start_month": min(mnths), "end_month": max(mnths),
                              "status": "❌ Not Started", "actuals": {}})
            tot = sum(a["weight"] for a in acts)
            if tot > 0 and abs(tot-100) > 1:
                for a in acts: a["weight"] = round(a["weight"]/tot*100, 2)
            return {"project_name": proj_name, "project_name_th": "", "start_date": start_date,
                    "end_date": "", "n_months": n_months, "total_budget": budget,
                    "contract_no": "", "project_owner": "NRCT",
                    "contractor": "Faculty of Engineering, CMU", "activities": acts}

        bulk_csv_lib = st.file_uploader("Upload WBS CSV", type=["csv"], key="lib_bulk_csv")
        if bulk_csv_lib:
            from collections import OrderedDict as _OD
            reader_lib = list(csv.DictReader(io.StringIO(bulk_csv_lib.read().decode("utf-8-sig"))))
            phase_map_lib = _OD()
            for r in reader_lib: phase_map_lib.setdefault(r.get("PHASE","Unknown"), []).append(r)
            st.markdown(f"**{len(reader_lib)} rows · {len(phase_map_lib)} phases**")
            prev_lib = [{"Phase": ph, "Activities": len(rows2), "Weight Sum": f"{sum(_wt(r.get('NOTES','')) for r in rows2):.0f}%"} for ph, rows2 in phase_map_lib.items()]
            st.dataframe(pd.DataFrame(prev_lib), use_container_width=True, hide_index=True)
            with st.form("lib_bulk_csv_form"):
                bc1, bc2, bc3 = st.columns(3)
                b_start = bc1.text_input("Start Date", "2025-10-01")
                b_nmo   = bc2.number_input("Months", 1, 36, 12)
                b_bud   = bc3.number_input("Budget/phase (THB)", value=20723000.0, step=100000.0)
                b_ovr   = st.checkbox("Overwrite existing", value=True)
                if st.form_submit_button("📦 Import All Phases", use_container_width=True, type="primary"):
                    imported = errors = 0
                    existing_keys = {e["key"] for e in db_list()}
                    for ph, rows2 in phase_map_lib.items():
                        try:
                            doc_code = _ext(rows2[0].get("NOTES",""),"Document") or re.sub(r'[^A-Za-z0-9_]','_', ph.split("/")[0].strip())[:10]
                            pname2   = rows2[0].get("PROJECT", ph)
                            rkey2    = re.sub(r'[^A-Za-z0-9_]','_', f"{doc_code}_{ph}")[:40]
                            if rkey2 in existing_keys and not b_ovr: continue
                            proj2 = csv_to_proj(rows2, pname2, b_start, int(b_nmo), float(b_bud))
                            proj2["_db_key"] = rkey2; proj2["_db_doc"] = doc_code; proj2["_db_phase"] = ph
                            db_save(rkey2, pname2, "", doc_code, ph, proj2)
                            db_log(rkey2, pname2, "📦 CSV Bulk Import", f"Phase: {ph}"); imported += 1
                        except Exception: errors += 1
                    st.success(f"✅ Imported **{imported}** phases | ⚠️ Errors: {errors}"); st.rerun()
        else:
            st.info("Upload `heymorning_task_import.csv`. Each PHASE becomes a separate project.")

    # ── Import single CSV/WBS ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("### 📄 Import Single CSV / WBS")
    st.markdown("Upload a CSV to parse, select a phase, and import into the active project.")

    def parse_wbs_csv_lib(file_bytes):
        text = file_bytes.decode("utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    def extract_notes_lib(notes, key):
        m = re.search(rf'{key}: ([^\|]+)', notes or ""); return m.group(1).strip() if m else ""
    def parse_months_lib(notes):
        raw = extract_notes_lib(notes, "Months")
        try: return [int(x) for x in raw.split(",") if x.strip()]
        except: return []
    def rows_to_acts_lib(rows, normalize=True):
        _acts = []
        for i, r in enumerate(rows):
            months = parse_months_lib(r.get("NOTES",""))
            if not months: continue
            try: wt = float(extract_notes_lib(r.get("NOTES",""), "Weight %"))
            except: wt = 0.0
            wbs = extract_notes_lib(r.get("NOTES",""), "WBS")
            try: pc = float(extract_notes_lib(r.get("NOTES",""), "Estimated cost"))
            except: pc = 0.0
            _acts.append({"no": wbs or str(i+1), "name": r.get("TASK NAME","").strip(),
                           "name_th": r.get("TASK NAME","").strip(), "weight": wt, "planned_cost": pc,
                           "start_month": min(months), "end_month": max(months),
                           "status": "❌ Not Started", "actuals": {}})
        if normalize and _acts:
            total = sum(a["weight"] for a in _acts)
            if total > 0 and abs(total-100.0) > 1.0:
                for a in _acts: a["weight"] = round(a["weight"]/total*100, 2)
        return _acts

    uploaded_lib = st.file_uploader("Upload WBS CSV file", type=["csv"], key="lib_csv_single")
    if uploaded_lib:
        raw_rows_lib = parse_wbs_csv_lib(uploaded_lib.read())
        st.success(f"✅ Loaded {len(raw_rows_lib)} rows")
        from collections import OrderedDict as _OD2
        pm_lib = _OD2()
        for r in raw_rows_lib: pm_lib.setdefault(r.get("PHASE","Unknown"), []).append(r)
        ph_rows_lib = [{"Phase": ph, "Activities": len(rows2),
                        "Weight Sum %": f"{sum(float(extract_notes_lib(r.get('NOTES',''),'Weight %') or 0) for r in rows2):.0f}",
                        "Date Range": f"{rows2[0].get('START DATE','')} → {rows2[-1].get('DUE DATE','')}"}
                       for ph, rows2 in pm_lib.items()]
        st.dataframe(pd.DataFrame(ph_rows_lib), use_container_width=True, hide_index=True)
        sel_ph_lib = st.selectbox("Select phase:", list(pm_lib.keys()), key="lib_sel_phase")
        acts_prev_lib = rows_to_acts_lib(pm_lib[sel_ph_lib])
        st.dataframe(pd.DataFrame([{"No": a["no"], "Task": a["name"][:50],
                                     "Wt%": f"{a['weight']:.1f}", "Cost": f"฿{a.get('planned_cost',0):,.0f}",
                                     "M Start": a["start_month"], "M End": a["end_month"]}
                                    for a in acts_prev_lib]), use_container_width=True, hide_index=True)
        with st.form("lib_single_import"):
            c1, c2 = st.columns(2)
            pn_lib   = c1.text_input("Project Name (EN)", f"Water Security — {sel_ph_lib}")
            pnth_lib = c2.text_input("Project Name (TH)", "โครงการขับเคลื่อนยุทธศาสตร์น้ำมั่นคง")
            c3,c4,c5 = st.columns(3)
            ps_lib   = c3.text_input("Start Date", pm_lib[sel_ph_lib][0].get("START DATE","2025-10-01")[:10])
            pe_lib   = c4.text_input("End Date",   pm_lib[sel_ph_lib][-1].get("DUE DATE","2026-09-30")[:10])
            nm_lib   = c5.number_input("Months", 1, 36, 12)
            bud_lib  = st.number_input("Total Budget (THB)", value=20723000.0, step=100000.0)
            ovr_lib  = st.checkbox("⚠️ Replace current project data", value=True)
            if st.form_submit_button("📥 Import & Save", use_container_width=True):
                nd = {"project_name": pn_lib, "project_name_th": pnth_lib,
                      "start_date": ps_lib, "end_date": pe_lib,
                      "n_months": int(nm_lib), "total_budget": float(bud_lib),
                      "contract_no": "", "project_owner": "NRCT",
                      "contractor": "Faculty of Engineering, CMU", "activities": acts_prev_lib}
                if ovr_lib:
                    save_data(nd); st.success(f"✅ Imported {len(acts_prev_lib)} activities!"); st.rerun()
                else:
                    cur = load_data(); existing_nos = {a["no"] for a in cur["activities"]}
                    added = 0
                    for a in acts_prev_lib:
                        if a["no"] not in existing_nos: cur["activities"].append(a); added += 1
                    save_data(cur); st.success(f"✅ Merged: {added} new activities."); st.rerun()
    else:
        st.info("Upload your WBS CSV file above to begin.")

    # ── Save current project to Library ──────────────────────────────────────
    st.markdown("---")
    st.markdown("#### ➕ Add Current Project to Library")
    with st.form("lib_save_to_db"):
        c1, c2, c3 = st.columns(3)
        lib_name  = c1.text_input("Project Name", data["project_name"])
        lib_doc   = c2.text_input("Document Code", "CUSTOM")
        lib_phase = c3.text_input("Phase / Tag", "Custom")
        lib_name_th = st.text_input("Project Name (TH)", data.get("project_name_th",""))
        if st.form_submit_button("💾 Save to Library", use_container_width=True):
            safe_key = re.sub(r'[^A-Za-z0-9_]','_', f"{lib_doc}_{lib_phase}")[:40]
            sp = dict(data); sp["project_name"] = lib_name; sp["project_name_th"] = lib_name_th
            sp["_db_key"] = safe_key; sp["_db_doc"] = lib_doc; sp["_db_phase"] = lib_phase
            db_save(safe_key, lib_name, lib_name_th, lib_doc, lib_phase, sp)
            db_log(safe_key, lib_name, "➕ Added to Library", f"{len(data.get('activities',[]))} activities")
            with open(DATA_FILE,"w",encoding="utf-8") as _f: json.dump(sp,_f,ensure_ascii=False,indent=2)
            st.cache_data.clear(); st.success(f"✅ Saved **{lib_name}** to library"); st.rerun()

    # ── Activity Log ──────────────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### 📋 Activity Log")
    log_filter_lib = st.selectbox("Filter", ["All projects"]+[e["project_name"] for e in db_list()],
                                   label_visibility="collapsed", key="lib_log_filter")
    log_key_lib = None
    if log_filter_lib != "All projects":
        matched = [e for e in db_list() if e["project_name"]==log_filter_lib]
        if matched: log_key_lib = matched[0]["key"]
    log_entries_lib = db_get_log(limit=60, db_key=log_key_lib)
    if not log_entries_lib:
        st.info("No activity recorded yet.")
    else:
        st.dataframe(pd.DataFrame([{"Timestamp": e["timestamp"], "Action": e["action"],
                                     "Project": e["project_name"][:45], "Detail": e["detail"]}
                                    for e in log_entries_lib]),
                     use_container_width=True, hide_index=True)
        lc1, lc2 = st.columns([4,1])
        lc1.caption(f"{len(log_entries_lib)} entries")
        if lc2.button("🗑️ Clear Log", use_container_width=True, key="lib_clear_log"):
            conn = db_connect()
            if log_key_lib: conn.execute("DELETE FROM save_log WHERE db_key=?", (log_key_lib,))
            else: conn.execute("DELETE FROM save_log")
            conn.commit(); conn.close(); st.rerun()

# ── PROJECT SETUP (merged) ────────────────────────────────────────────────────
elif page=="⚙️ Project Setup":
    st.markdown("# ⚙️ Project Setup")
    st.markdown("---")

    MANIFEST_PATH = "projects/project_manifest.json"
    PROJECTS_DIR  = "projects"

    tab1, tab4 = st.tabs([
        "📋 Project Info", "✏️ Activities"
    ])

    # ── TAB 1: Project Info ───────────────────────────────────────────────────
    with tab1:
        st.markdown("### 📋 Project Information")
        with st.form("pf"):
            c1,c2=st.columns(2)
            pn=c1.text_input("Project Name (EN)",data["project_name"])
            pt=c2.text_input("Project Name (TH)",data.get("project_name_th",""))
            c3,c4=st.columns(2)
            po=c3.text_input("Project Owner",data.get("project_owner",""))
            pc=c4.text_input("Contractor",data.get("contractor",""))
            c5,c6,c7=st.columns(3)
            ps=c5.text_input("Start Date",data["start_date"])
            pe=c6.text_input("End Date",data["end_date"])
            nm=c7.number_input("Months",1,36,data["n_months"])
            bg=st.number_input("Budget (THB)",value=float(data["total_budget"]),step=100000.0)
            cn=st.text_input("Contract No.",data.get("contract_no",""))
            if st.form_submit_button("💾 Save Project Info"):
                data.update({"project_name":pn,"project_name_th":pt,"project_owner":po,
                    "contractor":pc,"start_date":ps,"end_date":pe,"n_months":int(nm),
                    "total_budget":bg,"contract_no":cn})
                save_data(data); st.success("✅ Saved!"); st.rerun()

        if acts:
            st.markdown("---")
            total_cost = sum(a.get("planned_cost",0) for a in acts)
            st.caption(f"**{len(acts)} activities** | WBS cost: **฿{total_cost:,.0f}** | Budget: **฿{data.get('total_budget',0):,.0f}**")
            df=pd.DataFrame([{"No":a["no"],"Name":a["name"],"Wt%":a["weight"],
                "Planned Cost (฿)":a.get("planned_cost",0),
                "Start M":a["start_month"],"End M":a["end_month"],"Status":a.get("status","—")} for a in acts])
            st.dataframe(df,use_container_width=True,hide_index=True,
                         column_config={"Planned Cost (฿)": st.column_config.NumberColumn(format="฿%.0f")})

    with tab4:
        st.markdown("### ✏️ Manage Current Project Activities")
        st.markdown(f"**{data['project_name']}** — {len(acts)} activities loaded")
        st.markdown("Use the table below to **add new rows** (click ＋ at the bottom) or **delete rows** (select row → Delete key, or use the row checkbox). Click **💾 Save Changes** when done.")
        st.markdown("---")

        sopts_manage = ["Not Started","In Progress","Pending","Completed","Delayed","Cancelled"]

        # Build editable dataframe from current activities
        mrows = []
        for a in acts:
            sk = next((k for k in sopts_manage if k in a.get("status","")), sopts_manage[0])
            mrows.append({
                "Del":            False,
                "No":             str(a.get("no","") or ""),
                "Name (EN)":      str(a.get("name","") or ""),
                "Name (TH)":      str(a.get("name_th","") or ""),
                "Weight %":       float(a.get("weight",0)),
                "Planned Cost":   float(a.get("planned_cost",0)),
                "Start M":        int(a.get("start_month",1)),
                "End M":          int(a.get("end_month",1)),
                "Status":         sk,
            })

        mdf = pd.DataFrame(mrows) if mrows else pd.DataFrame(
            columns=["Del","No","Name (EN)","Name (TH)","Weight %","Planned Cost","Start M","End M","Status"])

        st.info("☑️ Check **Del** on any row then click **💾 Save Changes** to delete it. Click ＋ at the bottom to add new rows.", icon="ℹ️")

        managed = st.data_editor(
            mdf,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Del":          st.column_config.CheckboxColumn("🗑️ Del", width="small", help="Check to delete this row on save"),
                "No":           st.column_config.TextColumn("No", width="small", disabled=True),
                "Name (EN)":    st.column_config.TextColumn("Name (EN)", width="large"),
                "Name (TH)":    st.column_config.TextColumn("Name (TH)", width="medium"),
                "Weight %":     st.column_config.NumberColumn("Weight %", min_value=0.0, max_value=100.0, step=0.5, format="%.2f", width="small"),
                "Planned Cost": st.column_config.NumberColumn("Cost (฿)", min_value=0.0, step=10000.0, format="฿%.0f", width="medium"),
                "Start M":      st.column_config.NumberColumn("Start M", min_value=1, max_value=36, step=1, width="small"),
                "End M":        st.column_config.NumberColumn("End M",   min_value=1, max_value=36, step=1, width="small"),
                "Status":       st.column_config.SelectboxColumn("Status", width="medium", options=sopts_manage),
            },
            key="manage_editor"
        )

        wt_sum = managed["Weight %"].sum() if len(managed) > 0 else 0
        cost_sum = managed["Planned Cost"].sum() if len(managed) > 0 else 0
        c_info, c_save, c_norm = st.columns([3,1,1])
        c_info.markdown(
            f"**{len(managed)} activities** | Weight: **{wt_sum:.1f}%** {'✅' if abs(wt_sum-100)<1 else '⚠️ should be 100%'}"
            f" | Total cost: **฿{cost_sum:,.0f}**")

        if c_norm.button("⚖️ Normalize Weights", use_container_width=True):
            if wt_sum > 0:
                managed["Weight %"] = (managed["Weight %"] / wt_sum * 100).round(2)
                st.info("Weights normalized to 100%. Click 💾 Save Changes to apply.")

        # Show delete preview
        del_count = int(managed["Del"].sum()) if "Del" in managed.columns else 0
        if del_count > 0:
            st.warning(f"⚠️ {del_count} row(s) marked for deletion. Click **💾 Save Changes** to confirm.", icon="⚠️")

        if c_save.button("💾 Save Changes", use_container_width=True, type="primary"):
            new_acts = []
            for _, row in managed.iterrows():
                # Skip rows marked for deletion
                if row.get("Del", False): continue
                name = str(row.get("Name (EN)","")).strip()
                if not name: continue
                st_raw = str(row.get("Status","Not Started"))
                orig = next((a for a in acts if a.get("no")==str(row.get("No",""))), {})
                new_acts.append({
                    "no":          str(row.get("No","")),
                    "name":        name,
                    "name_th":     str(row.get("Name (TH)","")).strip() or name,
                    "weight":      float(row.get("Weight %", 0) or 0),
                    "planned_cost":float(row.get("Planned Cost", 0) or 0),
                    "start_month": int(row.get("Start M", 1) or 1),
                    "end_month":   int(row.get("End M", 1) or 1),
                    "status":      SMAP.get(st_raw, st_raw),
                    "actuals":     orig.get("actuals", {}),
                })
            data["activities"] = new_acts
            save_data(data)
            st.success(f"✅ Saved {len(new_acts)} activities!"); st.rerun()


# ── PROCESS GUIDE ─────────────────────────────────────────────────────────────
elif page=="📋 Process Guide":
    st.markdown("# 📋 Process Guide")
    st.markdown("Complete pipeline — from project proposal document to live S-Curve dashboard.")
    st.markdown("---")

    # ── PIPELINE OVERVIEW ────────────────────────────────────────────────────
    st.markdown("## 🔄 Full Pipeline Overview")
    st.markdown("""
<div style="background:#0A0A0A;border-radius:16px;padding:22px 26px;font-family:monospace;font-size:.9rem;line-height:2.4">
<span style="color:#1AE06B;font-weight:700">STEP 1</span> &nbsp;📄 <span style="color:#FFFFFF">Project Proposal Document</span> &nbsp;<span style="color:#6B6B6B">(.docx / .pdf)</span><br>
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<span style="color:#1AE06B">↓</span><br>
<span style="color:#1AE06B;font-weight:700">STEP 2</span> &nbsp;🤖 <span style="color:#FFFFFF">AI Extraction</span> &nbsp;<span style="color:#6B6B6B">(Claude / ChatGPT + WBS Prompt)</span><br>
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<span style="color:#1AE06B">↓</span><br>
<span style="color:#1AE06B;font-weight:700">STEP 3</span> &nbsp;📊 <span style="color:#FFFFFF">heymorning_task_import.csv</span> &nbsp;<span style="color:#6B6B6B">(WBS data table)</span><br>
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<span style="color:#1AE06B">↓</span><br>
<span style="color:#1AE06B;font-weight:700">STEP 4</span> &nbsp;⚙️ <span style="color:#FFFFFF">wbs_extractor.py</span> &nbsp;<span style="color:#6B6B6B">(auto-split by sub-project)</span><br>
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<span style="color:#1AE06B">↓</span><br>
<span style="color:#1AE06B;font-weight:700">STEP 5</span> &nbsp;📁 <span style="color:#FFFFFF">projects/*.json</span> &nbsp;<span style="color:#6B6B6B">(20 pre-built project files)</span><br>
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<span style="color:#1AE06B">↓</span><br>
<span style="color:#1AE06B;font-weight:700">STEP 6</span> &nbsp;📥 <span style="color:#FFFFFF">Import WBS → Project Library</span> &nbsp;<span style="color:#6B6B6B">(one-click load)</span><br>
&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;<span style="color:#1AE06B">↓</span><br>
<span style="color:#1AE06B;font-weight:700">STEP 7</span> &nbsp;📈 <span style="color:#1AE06B;font-weight:700">Dashboard / S-Curve / Gantt / EVM</span> &nbsp;✅ Live
</div>
""", unsafe_allow_html=True)

    st.markdown("---")

    # ── STEP CARDS ───────────────────────────────────────────────────────────
    steps = [
        ("1","📄","Prepare Project Proposal",
         "Have your project proposal ready as a Word (.docx) or PDF file. The document should contain schedule tables (monthly Gantt), budget breakdown, and activity lists.",
         ["Schedule tables with month columns (fiscal months 1–12)",
          "Budget table with activity weights or costs",
          "Work package / milestone descriptions"]),
        ("2","🤖","Run AI Extraction (WBS Prompt)",
         "Open a new Claude or ChatGPT conversation. Copy the prompt from `WBS_Extraction_Prompt.md` (in your GitHub repo), paste it, and attach your proposal document.",
         ["Open: claude.ai or chatgpt.com",
          "Copy full prompt from WBS_Extraction_Prompt.md",
          "Attach proposal document",
          "Send → AI outputs WBS table + CSV"]),
        ("3","💾","Save Output as CSV",
         "The AI will output a CSV block. Copy it and save as `heymorning_task_import.csv`. Each row = one WBS activity with months, weight, cost, and WBS code in the NOTES column.",
         ["Copy the ```csv block from AI response",
          "Paste into Notepad / VS Code",
          'Save as: heymorning_task_import.csv (UTF-8 encoding)',
          "Verify: header row must match expected columns"]),
        ("4","⚙️","Auto-Split by Sub-Project (Optional)",
         "Run `wbs_extractor.py` to automatically separate the CSV into one JSON per sub-project/province. This regenerates all files in the `projects/` folder.",
         ["Open Terminal / Command Prompt",
          "Navigate to scurve_webapp/ folder",
          "Run: python wbs_extractor.py --csv heymorning_task_import.csv --outdir projects --combine-doc",
          "Push updated JSONs to GitHub"]),
        ("5","📥","Import into the App",
         "Use the Project Setup page (sidebar) to load your project data. Two options: WBS Library (pre-built) or Import CSV (manual).",
         ["Sidebar → ⚙️ Project Setup",
          "Tab 1: Project Library → click ▶ Load next to your project",
          "OR Tab 2: Upload CSV → select phase → Import & Save",
          "Dashboard updates automatically"]),
        ("6","📝","Update Monthly Progress",
         "Each month, open Update Progress to enter actual completion % per activity and update status. Then commit project_data.json to GitHub to persist the changes.",
         ["Sidebar → 📝 Update Progress",
          "Expand each activity panel",
          "Enter actual % for the current month",
          "Click 💾 Save All Activities",
          "Commit project_data.json to GitHub to save permanently"]),
    ]

    for no, icon, title, desc, bullets in steps:
        with st.expander(f"**Step {no} — {icon} {title}**", expanded=(no=="1")):
            st.markdown(f"<p style='color:#1F2937'>{desc}</p>", unsafe_allow_html=True)
            for b in bullets:
                st.markdown(f"<div style='color:#15803D;padding:2px 0 2px 16px'>✦ {b}</div>",
                            unsafe_allow_html=True)
            st.markdown("")

    st.markdown("---")

    # ── WBS PROMPT DISPLAY ────────────────────────────────────────────────────
    st.markdown("## 🤖 AI Extraction Prompt")

    PROMPT_PATH = "WBS_Extraction_Prompt.md"
    if os.path.exists(PROMPT_PATH):
        with open(PROMPT_PATH, encoding="utf-8") as _pf:
            prompt_text = _pf.read()
        if "=== PASTE PROMPT BELOW ===" in prompt_text:
            prompt_section = prompt_text.split("=== PASTE PROMPT BELOW ===")[1].split("=== END OF PROMPT ===")[0].strip()
        else:
            prompt_section = prompt_text

        # Instruction banner + action buttons
        st.markdown("""
<div style="background:#FFFFFF;border:1.5px solid #E0E0E0;border-radius:14px;padding:14px 20px;margin-bottom:12px">
<b style="color:#0A0A0A;font-size:1.05rem">📋 How to use this prompt</b><br>
<span style="color:#3B3B3B;font-size:.9rem">
1. Click <b>⬇️ Download Prompt</b> (save to your computer) <b>OR</b> select all text below and copy<br>
2. Open <b>claude.ai</b> or <b>chatgpt.com</b> → start a new conversation<br>
3. Paste the prompt, then <b>attach your project proposal (.docx / .pdf)</b><br>
4. The AI outputs a CSV block → save as <code>heymorning_task_import.csv</code><br>
5. Go to <b>⚙️ Project Setup → Import CSV/WBS</b> to load it into the app
</span>
</div>
""", unsafe_allow_html=True)

        # Action buttons row
        btn_col1, btn_col2, btn_col3 = st.columns([2, 2, 4])

        # Download button — saves prompt as .txt file
        btn_col1.download_button(
            label="⬇️ Download Prompt (.txt)",
            data=prompt_section.encode("utf-8"),
            file_name="WBS_Extraction_Prompt.txt",
            mime="text/plain",
            use_container_width=True,
        )

        # Download as .md file
        btn_col2.download_button(
            label="📄 Download Prompt (.md)",
            data=prompt_section.encode("utf-8"),
            file_name="WBS_Extraction_Prompt.md",
            mime="text/markdown",
            use_container_width=True,
        )

        btn_col3.info("💡 Or click inside the text box below → **Ctrl+A** → **Ctrl+C** to copy all")

        # Full prompt in a tall text area — easy to select-all and copy
        st.text_area(
            label="Full AI Extraction Prompt (select all → copy)",
            value=prompt_section,
            height=520,
            key="prompt_textarea",
            help="Click inside → Ctrl+A to select all → Ctrl+C to copy"
        )

        # Also show as code block for syntax highlighting + built-in copy icon
        with st.expander("📋 View as formatted code (with copy icon ↗)", expanded=False):
            st.code(prompt_section, language="markdown")

    else:
        st.warning("WBS_Extraction_Prompt.md not found. Make sure it is in the same folder as app.py.")
        st.markdown("**Expected location:** `scurve_webapp/WBS_Extraction_Prompt.md`")

    st.markdown("---")

    # ── NOTES FIELD FORMAT ────────────────────────────────────────────────────
    st.markdown("## 📌 Required CSV Format")
    st.markdown("The NOTES column must contain this exact format for the importer to parse correctly:")
    st.code("WBS: FULL.1.01 | Document: FULL | FY(BE): 2569 | Months: 1,2,3 | Weight %: 10.0 | Estimated cost: 2072300.0 | Duration days: 92 | Source: table 19, row 2", language="text")
    st.markdown("**Full CSV header:**")
    st.code("PROJECT,PHASE,TASK NAME,DESCRIPTION,IMPORTANT,URGENT,STATUS,PRIORITY,ASSIGNED TO,START DATE,DUE DATE,KANBAN STAGE,PROGRESS,NOTES", language="text")

    st.markdown("---")

    # ── FISCAL MONTH TABLE ────────────────────────────────────────────────────
    st.markdown("## 📅 Fiscal Month Reference (FY 2569)")
    fm_data = {
        "Fiscal Month": list(range(1,13)),
        "Calendar Month": ["Oct","Nov","Dec","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep"],
        "Date (FY2569)": ["2025-10","2025-11","2025-12","2026-01","2026-02","2026-03",
                           "2026-04","2026-05","2026-06","2026-07","2026-08","2026-09"],
    }
    st.dataframe(pd.DataFrame(fm_data), use_container_width=True, hide_index=True)
